#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import datetime
from database import db
from auth import auth
from printer import printer
import subprocess
import os

class RestaurantPOS:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Sistema POS Restaurante")
        self.root.geometry("1024x768")
        self.root.configure(bg='#f5f5f5')
        
        # Configurar estilo
        self.setup_styles()
        
        # Variables
        self.orden_actual = {'items': [], 'total': 0.0}
        self.cliente_actual = None
        
        # Inicializar
        if not db.connect():
            messagebox.showerror("Error", "No se pudo conectar a la base de datos")
            self.root.quit()
            return
        
        # Mostrar login
        self.show_login()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar colores - esquema azul consistente
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), background='#f8f9fa', foreground='#2c3e50')
        style.configure('Header.TLabel', font=('Arial', 12, 'bold'), background='#f8f9fa', foreground='#34495e')
        style.configure('Blue.TButton', foreground='white', background='#3498db')
        style.configure('DarkBlue.TButton', foreground='white', background='#2980b9')
        style.configure('Gray.TButton', foreground='white', background='#95a5a6')
        style.configure('LightBlue.TButton', foreground='white', background='#5dade2')

    def show_login(self):
        self.clear_window()
        
        # Frame principal centrado
        main_frame = tk.Frame(self.root, bg='#f5f5f5')
        main_frame.pack(expand=True, fill='both')
        
        # Frame de login centrado
        login_frame = tk.Frame(main_frame, bg='white', relief='raised', bd=2)
        login_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=300)
        
        # Título
        title_label = tk.Label(login_frame, text="Sistema POS Restaurante", 
                              font=('Arial', 18, 'bold'), bg='white', fg='#2c3e50')
        title_label.pack(pady=20)
        
        # Campos de login
        tk.Label(login_frame, text="Usuario:", font=('Arial', 12), bg='white').pack(pady=5)
        self.username_entry = tk.Entry(login_frame, font=('Arial', 12), width=20)
        self.username_entry.pack(pady=5)
        
        tk.Label(login_frame, text="Contraseña:", font=('Arial', 12), bg='white').pack(pady=5)
        self.password_entry = tk.Entry(login_frame, font=('Arial', 12), width=20, show='*')
        self.password_entry.pack(pady=5)
        
        # Botón de login
        login_btn = tk.Button(login_frame, text="Iniciar Sesión", font=('Arial', 12, 'bold'),
                             bg='#3498db', fg='white', width=15, command=self.login)
        login_btn.pack(pady=20)
        
        # Bind Enter key
        self.username_entry.bind('<Return>', lambda e: self.password_entry.focus())
        self.password_entry.bind('<Return>', lambda e: self.login())
        
        self.username_entry.focus()

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        if not username or not password:
            messagebox.showerror("Error", "Por favor complete todos los campos")
            return
        
        if auth.login(username, password):
            if auth.current_user['tipo'] == 'cajero':
                # Verificar si necesita iniciar turno
                if not auth.current_turno:
                    self.show_iniciar_turno()
                else:
                    self.show_cajero_menu()
            elif auth.current_user['tipo'] in ['administrador', 'superusuario']:
                self.show_admin_menu()
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")

    def show_iniciar_turno(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Iniciar Turno")
        dialog.geometry("380x280")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal compacto
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="🕐 INICIAR NUEVO TURNO", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        # Frame para fondo inicial - más compacto
        fondo_frame = tk.LabelFrame(main_frame, text="Fondo Inicial de Caja", font=('Arial', 11),
                                   bg='#f8f9fa', fg='#2c3e50')
        fondo_frame.pack(pady=(0,15), fill='x')
        
        entry_frame = tk.Frame(fondo_frame, bg='#f8f9fa')
        entry_frame.pack(pady=10)
        
        tk.Label(entry_frame, text="$", font=('Arial', 16, 'bold'), bg='#f8f9fa').pack(side='left')
        fondo_entry = tk.Entry(entry_frame, font=('Arial', 16), width=8, justify='center',
                              relief='solid', bd=1)
        fondo_entry.pack(side='left', padx=5)
        
        def crear_turno():
            try:
                fondo_text = fondo_entry.get().strip()
                if not fondo_text:
                    messagebox.showerror("Error", "Por favor ingrese el fondo inicial")
                    fondo_entry.focus()
                    return
                
                fondo = float(fondo_text)
                if fondo < 0:
                    messagebox.showerror("Error", "El fondo inicial no puede ser negativo")
                    fondo_entry.focus()
                    return
                
                if auth.crear_turno(fondo):
                    messagebox.showinfo("Éxito", f"Turno iniciado correctamente\nFondo inicial: ${fondo:.2f}")
                    dialog.destroy()
                    self.show_cajero_menu()
                else:
                    messagebox.showerror("Error", "No se pudo crear el turno")
            except ValueError:
                messagebox.showerror("Error", "Por favor ingrese un monto válido")
                fondo_entry.focus()
        
        # Botones - más compactos
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(pady=(0,5), fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=self.root.quit).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="INICIAR TURNO", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=crear_turno).pack(side='right', padx=5)
        
        fondo_entry.focus()
        fondo_entry.bind('<Return>', lambda e: crear_turno())

    def show_cajero_menu(self):
        self.clear_window()
        
        # Header
        header_frame = tk.Frame(self.root, bg='#34495e', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        # Información del usuario y turno
        user_info = f"Usuario: {auth.current_user['nombre']} | Turno: {auth.current_turno['id']}"
        tk.Label(header_frame, text=user_info, font=('Arial', 12, 'bold'), 
                bg='#34495e', fg='white').pack(side='left', padx=20, pady=20)
        
        # Botones del header
        btn_frame = tk.Frame(header_frame, bg='#34495e')
        btn_frame.pack(side='right', padx=20, pady=10)
        
        tk.Button(btn_frame, text="Historial", command=self.show_historial).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Arqueo", command=self.show_arqueo).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Corte Caja", command=self.show_corte_caja).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Cerrar Sesión", command=self.logout).pack(side='left', padx=5)
        
        # Contenido principal
        main_frame = tk.Frame(self.root, bg='#f5f5f5')
        main_frame.pack(expand=True, fill='both', padx=20, pady=20)
        
        # Botones de tipo de orden
        orden_frame = tk.LabelFrame(main_frame, text="Crear Nueva Orden", font=('Arial', 12, 'bold'))
        orden_frame.pack(fill='x', pady=(0, 20))
        
        btn_frame = tk.Frame(orden_frame)
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Venta Mostrador", font=('Arial', 12, 'bold'),
                 bg='#3498db', fg='white', width=15, height=2,
                 command=lambda: self.nueva_orden('mostrador')).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="Pedido Restaurante", font=('Arial', 12, 'bold'),
                 bg='#2980b9', fg='white', width=15, height=2,
                 command=lambda: self.nueva_orden('restaurante')).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="Pedido Telefónico", font=('Arial', 12, 'bold'),
                 bg='#5dade2', fg='white', width=15, height=2,
                 command=lambda: self.nueva_orden('telefonico')).pack(side='left', padx=10)

    def nueva_orden(self, tipo_orden):
        self.tipo_orden = tipo_orden
        self.orden_actual = {'items': [], 'total': 0.0}
        self.cliente_actual = None
        
        if tipo_orden == 'telefonico':
            self.seleccionar_cliente()
        else:
            self.show_orden_screen()

    def seleccionar_cliente(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Seleccionar Cliente")
        dialog.geometry("420x300")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal compacto
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="📞 BUSCAR CLIENTE", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        # Búsqueda por teléfono - más compacto
        search_frame = tk.LabelFrame(main_frame, text="Buscar por Teléfono", font=('Arial', 11),
                                    bg='#f8f9fa', fg='#2c3e50')
        search_frame.pack(pady=(0,15), fill='x')
        
        entry_frame = tk.Frame(search_frame, bg='#f8f9fa')
        entry_frame.pack(pady=10)
        
        tk.Label(entry_frame, text="📱", font=('Arial', 16), bg='#f8f9fa').pack(side='left', padx=5)
        phone_entry = tk.Entry(entry_frame, font=('Arial', 12), width=15, relief='solid', bd=1)
        phone_entry.pack(side='left', padx=5, fill='x', expand=True)
        
        def buscar_cliente():
            telefono = phone_entry.get().strip()
            if not telefono:
                messagebox.showerror("Error", "Por favor ingrese un número telefónico")
                phone_entry.focus()
                return
            
            query = "SELECT * FROM clientes WHERE telefono = %s"
            cliente = db.execute_one(query, (telefono,))
            
            if cliente:
                self.cliente_actual = cliente
                messagebox.showinfo("Cliente encontrado", 
                                  f"Cliente: {cliente['nombre']}\nTeléfono: {cliente['telefono']}\nDirección: {cliente['direccion'] or 'No especificada'}")
                dialog.destroy()
                self.show_orden_screen()
            else:
                response = messagebox.askyesno("Cliente no encontrado", 
                                             f"No se encontró un cliente con el teléfono {telefono}\n\n¿Desea crear un nuevo cliente?")
                if response:
                    self.crear_cliente_dialog(telefono, dialog)
        
        # Botones - más compactos
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(pady=(0,5), fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="BUSCAR CLIENTE", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=15, height=1,
                 command=buscar_cliente).pack(side='right', padx=5)
        
        phone_entry.focus()
        phone_entry.bind('<Return>', lambda e: buscar_cliente())

    def crear_cliente_dialog(self, telefono, parent_dialog):
        dialog = tk.Toplevel(self.root)
        dialog.title("Crear Nuevo Cliente")
        dialog.geometry("400x320")
        dialog.resizable(False, False)
        dialog.transient(parent_dialog)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal compacto
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="👤 CREAR NUEVO CLIENTE", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        # Campos del formulario - más compactos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(pady=(0,15), fill='x')
        
        # Nombre
        tk.Label(fields_frame, text="Nombre completo:", font=('Arial', 11), 
                bg='#f8f9fa', fg='#2c3e50').grid(row=0, column=0, sticky='w', pady=6)
        nombre_entry = tk.Entry(fields_frame, width=20, font=('Arial', 11), relief='solid', bd=1)
        nombre_entry.grid(row=0, column=1, pady=6, padx=8, sticky='ew')
        
        # Teléfono
        tk.Label(fields_frame, text="Teléfono:", font=('Arial', 11), 
                bg='#f8f9fa', fg='#2c3e50').grid(row=1, column=0, sticky='w', pady=6)
        telefono_entry = tk.Entry(fields_frame, width=20, font=('Arial', 11), relief='solid', bd=1)
        telefono_entry.insert(0, telefono)
        telefono_entry.grid(row=1, column=1, pady=6, padx=8, sticky='ew')
        
        # Dirección
        tk.Label(fields_frame, text="Dirección:", font=('Arial', 11), 
                bg='#f8f9fa', fg='#2c3e50').grid(row=2, column=0, sticky='w', pady=6)
        direccion_entry = tk.Entry(fields_frame, width=20, font=('Arial', 11), relief='solid', bd=1)
        direccion_entry.grid(row=2, column=1, pady=6, padx=8, sticky='ew')
        
        # Configurar grid
        fields_frame.grid_columnconfigure(1, weight=1)
        
        def guardar_cliente():
            nombre = nombre_entry.get().strip()
            telefono = telefono_entry.get().strip()
            direccion = direccion_entry.get().strip()
            
            if not nombre:
                messagebox.showerror("Error", "El nombre es obligatorio")
                nombre_entry.focus()
                return
            
            if not telefono:
                messagebox.showerror("Error", "El teléfono es obligatorio")
                telefono_entry.focus()
                return
            
            query = """
            INSERT INTO clientes (nombre, telefono, direccion)
            VALUES (%s, %s, %s)
            """
            cliente_id = db.execute_one(query, (nombre, telefono, direccion))
            
            if cliente_id:
                query = "SELECT * FROM clientes WHERE id = %s"
                self.cliente_actual = db.execute_one(query, (cliente_id,))
                messagebox.showinfo("Éxito", f"Cliente '{nombre}' creado correctamente")
                dialog.destroy()
                parent_dialog.destroy()
                self.show_orden_screen()
            else:
                messagebox.showerror("Error", "No se pudo crear el cliente. Verifique que el teléfono no esté duplicado.")
        
        # Botones - más compactos
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(pady=(0,5), fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="GUARDAR CLIENTE", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=15, height=1,
                 command=guardar_cliente).pack(side='right', padx=5)
        
        nombre_entry.focus()
        nombre_entry.bind('<Return>', lambda e: telefono_entry.focus())
        telefono_entry.bind('<Return>', lambda e: direccion_entry.focus())
        direccion_entry.bind('<Return>', lambda e: guardar_cliente())

    def show_orden_screen(self):
        self.clear_window()
        
        # Header con información de la orden
        header_frame = tk.Frame(self.root, bg='#34495e', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        info_text = f"Tipo: {self.tipo_orden.title()}"
        if self.cliente_actual:
            info_text += f" | Cliente: {self.cliente_actual['nombre']}"
        
        tk.Label(header_frame, text=info_text, font=('Arial', 12, 'bold'), 
                bg='#34495e', fg='white').pack(side='left', padx=20, pady=20)
        
        tk.Button(header_frame, text="Volver", command=self.show_cajero_menu).pack(side='right', padx=20, pady=15)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f5f5f5')
        main_frame.pack(expand=True, fill='both', padx=20, pady=20)
        
        # Frame izquierdo - Menu de productos
        left_frame = tk.Frame(main_frame)
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        # Obtener productos
        self.show_productos_menu(left_frame)
        
        # Frame derecho - Resumen de orden
        right_frame = tk.Frame(main_frame, width=300)
        right_frame.pack(side='right', fill='y')
        right_frame.pack_propagate(False)
        
        self.show_orden_summary(right_frame)

    def show_productos_menu(self, parent):
        # Título
        tk.Label(parent, text="Seleccionar Productos", font=('Arial', 14, 'bold')).pack(pady=10)
        
        # Obtener productos por categoría
        query = """
        SELECT c.nombre as categoria, p.id, p.nombre, p.precio
        FROM productos p
        JOIN categorias c ON p.categoria_id = c.id
        WHERE p.activo = 1
        ORDER BY c.nombre, p.nombre
        """
        productos = db.execute_query(query)
        
        # Crear notebook para categorías
        notebook = ttk.Notebook(parent)
        notebook.pack(fill='both', expand=True, pady=10)
        
        # Agrupar productos por categoría
        categorias = {}
        for producto in productos:
            cat = producto['categoria']
            if cat not in categorias:
                categorias[cat] = []
            categorias[cat].append(producto)
        
        # Crear tabs por categoría
        for categoria, productos_cat in categorias.items():
            tab_frame = ttk.Frame(notebook)
            notebook.add(tab_frame, text=categoria)
            
            # Crear grid de productos
            products_frame = tk.Frame(tab_frame)
            products_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            row, col = 0, 0
            for producto in productos_cat:
                btn_frame = tk.Frame(products_frame, relief='raised', bd=1, bg='white')
                btn_frame.grid(row=row, column=col, padx=5, pady=5, sticky='nsew')
                
                # Configurar grid
                products_frame.grid_columnconfigure(col, weight=1, minsize=150)
                products_frame.grid_rowconfigure(row, weight=1, minsize=80)
                
                # Información del producto
                tk.Label(btn_frame, text=producto['nombre'], font=('Arial', 11, 'bold'), 
                        bg='white').pack(pady=5)
                tk.Label(btn_frame, text=f"${producto['precio']:.2f}", font=('Arial', 10), 
                        bg='white', fg='#27ae60').pack()
                
                # Hacer clickeable
                def add_product(p=producto):
                    self.agregar_producto(p)
                
                btn_frame.bind("<Button-1>", lambda e, p=producto: self.agregar_producto(p))
                for widget in btn_frame.winfo_children():
                    widget.bind("<Button-1>", lambda e, p=producto: self.agregar_producto(p))
                
                col += 1
                if col >= 4:  # 4 productos por fila
                    col = 0
                    row += 1

    def show_orden_summary(self, parent):
        # Título
        tk.Label(parent, text="Resumen de Orden", font=('Arial', 14, 'bold')).pack(pady=10)
        
        # Frame para nombre cliente (solo mostrador)
        if self.tipo_orden == 'mostrador':
            cliente_frame = tk.Frame(parent)
            cliente_frame.pack(fill='x', padx=10, pady=5)
            
            tk.Label(cliente_frame, text="Nombre cliente:").pack(anchor='w')
            self.cliente_nombre_entry = tk.Entry(cliente_frame)
            self.cliente_nombre_entry.pack(fill='x', pady=2)
        
        # Lista de productos
        list_frame = tk.Frame(parent)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Scrollable listbox
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.items_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, height=15)
        self.items_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.items_listbox.yview)
        
        # Frame para total
        total_frame = tk.Frame(parent)
        total_frame.pack(fill='x', padx=10, pady=10)
        
        self.total_label = tk.Label(total_frame, text="TOTAL: $0.00", 
                                   font=('Arial', 16, 'bold'), fg='#27ae60')
        self.total_label.pack(anchor='center')
        
        # Botones de acción
        btn_frame = tk.Frame(parent)
        btn_frame.pack(fill='x', padx=10, pady=20)
        
        tk.Button(btn_frame, text="Eliminar Item", bg='#95a5a6', fg='white',
                 command=self.eliminar_item).pack(fill='x', pady=2)
        
        tk.Button(btn_frame, text="Procesar Pago", bg='#3498db', fg='white',
                 font=('Arial', 12, 'bold'), command=self.procesar_pago).pack(fill='x', pady=10)
        
        self.actualizar_resumen()

    def agregar_producto(self, producto):
        # Buscar si ya existe en la orden
        for item in self.orden_actual['items']:
            if item['id'] == producto['id']:
                item['cantidad'] += 1
                # Asegurar que el precio sea float
                precio_float = float(producto['precio'])
                item['subtotal'] = item['cantidad'] * precio_float
                break
        else:
            # Agregar nuevo item - convertir precio a float
            precio_float = float(producto['precio'])
            self.orden_actual['items'].append({
                'id': producto['id'],
                'nombre': producto['nombre'],
                'precio': precio_float,
                'cantidad': 1,
                'subtotal': precio_float
            })
        
        self.actualizar_resumen()

    def eliminar_item(self):
        selection = self.items_listbox.curselection()
        if selection:
            index = selection[0]
            if index < len(self.orden_actual['items']):
                del self.orden_actual['items'][index]
                self.actualizar_resumen()

    def actualizar_resumen(self):
        # Limpiar listbox
        self.items_listbox.delete(0, tk.END)
        
        # Agregar items
        total = 0
        for item in self.orden_actual['items']:
            text = f"{item['cantidad']}x {item['nombre']} - ${item['subtotal']:.2f}"
            self.items_listbox.insert(tk.END, text)
            total += item['subtotal']
        
        # Actualizar total
        self.orden_actual['total'] = total
        self.total_label.config(text=f"TOTAL: ${total:.2f}")

    def procesar_pago(self):
        print("DEBUG: Iniciando procesar_pago")
        print(f"DEBUG: Items en orden: {len(self.orden_actual['items'])}")
        print(f"DEBUG: Total: {self.orden_actual['total']}")
        print(f"DEBUG: Tipo de orden: {self.tipo_orden}")
        
        if not self.orden_actual['items']:
            messagebox.showerror("Error", "La orden está vacía")
            return
        
        # Validar nombre para mostrador
        cliente_nombre = None
        if self.tipo_orden == 'mostrador':
            cliente_nombre = self.cliente_nombre_entry.get().strip()
            print(f"DEBUG: Nombre cliente mostrador: '{cliente_nombre}'")
            if not cliente_nombre:
                messagebox.showerror("Error", "Por favor ingrese el nombre del cliente")
                return
        
        print(f"DEBUG: Llamando show_payment_dialog con cliente_nombre: {cliente_nombre}")
        self.show_payment_dialog(cliente_nombre)

    def show_payment_dialog(self, cliente_nombre=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("Procesar Pago")
        dialog.geometry("420x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal compacto
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Total a pagar - más compacto
        tk.Label(main_frame, text="TOTAL A PAGAR", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,5))
        tk.Label(main_frame, text=f"${self.orden_actual['total']:.2f}", 
                font=('Arial', 24, 'bold'), fg='#3498db', bg='#f8f9fa').pack(pady=(0,15))
        
        # Método de pago - más compacto
        payment_frame = tk.LabelFrame(main_frame, text="Método de Pago", font=('Arial', 11), 
                                     bg='#f8f9fa', fg='#2c3e50')
        payment_frame.pack(pady=(0,10), fill='x')
        
        payment_method = tk.StringVar(value='efectivo')
        
        radio_frame = tk.Frame(payment_frame, bg='#f8f9fa')
        radio_frame.pack(pady=8)
        
        tk.Radiobutton(radio_frame, text="💵 Efectivo", variable=payment_method, 
                      value='efectivo', font=('Arial', 11), bg='#f8f9fa').pack(side='left', padx=15)
        tk.Radiobutton(radio_frame, text="💳 Tarjeta", variable=payment_method, 
                      value='tarjeta', font=('Arial', 11), bg='#f8f9fa').pack(side='left', padx=15)
        
        # Monto pagado - más compacto
        amount_frame = tk.LabelFrame(main_frame, text="Monto Pagado", font=('Arial', 11), 
                                    bg='#f8f9fa', fg='#2c3e50')
        amount_frame.pack(pady=(0,10), fill='x')
        
        entry_frame = tk.Frame(amount_frame, bg='#f8f9fa')
        entry_frame.pack(pady=8)
        
        tk.Label(entry_frame, text="$", font=('Arial', 16, 'bold'), bg='#f8f9fa').pack(side='left')
        amount_entry = tk.Entry(entry_frame, font=('Arial', 16), width=8, justify='center',
                               relief='solid', bd=1)
        amount_entry.pack(side='left', padx=5)
        amount_entry.focus()
        
        # Cambio - más compacto
        change_frame = tk.Frame(main_frame, bg='#e3f2fd', relief='solid', bd=1)
        change_frame.pack(pady=(0,15), fill='x')
        
        change_label = tk.Label(change_frame, text="Ingrese el monto pagado", 
                               font=('Arial', 12, 'bold'), fg='#3498db', bg='#e3f2fd')
        change_label.pack(pady=8)
        
        def calculate_change():
            try:
                amount_text = amount_entry.get().strip()
                if amount_text:
                    amount = float(amount_text)
                    # Convertir total a float para evitar problemas de tipo
                    total_orden = float(self.orden_actual['total'])
                    change = amount - total_orden
                    if change >= 0:
                        change_label.config(text=f"CAMBIO: ${change:.2f}", fg='#27ae60')
                    elif change >= -0.01:
                        change_label.config(text="MONTO EXACTO", fg='#27ae60')
                    else:
                        change_label.config(text="❌ MONTO INSUFICIENTE", fg='#e74c3c')
                else:
                    change_label.config(text="Ingrese el monto pagado", fg='#3498db')
            except ValueError:
                change_label.config(text="❌ MONTO INVÁLIDO", fg='#e74c3c')
            except Exception as e:
                print(f"ERROR en calculate_change: {e}")
                change_label.config(text="❌ ERROR", fg='#e74c3c')
        
        amount_entry.bind('<KeyRelease>', lambda e: calculate_change())
        
        def confirmar_pago():
            try:
                amount_text = amount_entry.get().strip()
                if not amount_text:
                    messagebox.showerror("Error", "Por favor ingrese el monto pagado")
                    amount_entry.focus()
                    return
                
                amount = float(amount_text)
                # Convertir total a float para la comparación
                total_orden = float(self.orden_actual['total'])
                
                if amount < total_orden - 0.01:
                    messagebox.showerror("Error", "Monto insuficiente")
                    amount_entry.focus()
                    return
                
                # Usar el parámetro cliente_nombre que se pasó a la función show_payment_dialog
                self.crear_orden_db(payment_method.get(), amount, cliente_nombre)
                dialog.destroy()
                
            except ValueError:
                messagebox.showerror("Error", "Por favor ingrese un monto válido")
                amount_entry.focus()
            except Exception as e:
                messagebox.showerror("Error", f"Error procesando el pago: {str(e)}")
                print(f"Error en confirmar_pago: {e}")  # Para debug
        
        # Botones - más compactos
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(pady=(0,5), fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="CONFIRMAR PAGO", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=15, height=1,
                 command=confirmar_pago).pack(side='right', padx=5)
        
        amount_entry.bind('<Return>', lambda e: confirmar_pago())

    def crear_orden_db(self, metodo_pago, monto_pagado, cliente_nombre=None):
        try:
            print(f"DEBUG: Creando orden - metodo_pago: {metodo_pago}, monto_pagado: {monto_pagado}, cliente_nombre: {cliente_nombre}")
            print(f"DEBUG: Items en orden: {len(self.orden_actual['items'])}")
            print(f"DEBUG: Total orden: {self.orden_actual['total']}")
            
            # Verificar que haya items en la orden
            if not self.orden_actual['items']:
                raise Exception("La orden está vacía")
            
            # Verificar que hay un turno activo
            if not auth.current_turno:
                raise Exception("No hay turno activo")
            
            # Generar número de orden
            import time
            numero_orden = f"ORD{int(time.time())}"
            print(f"DEBUG: Número de orden generado: {numero_orden}")
            
            # Convertir a float para evitar problemas de tipo Decimal
            total_orden = float(self.orden_actual['total'])
            monto_pagado_float = float(monto_pagado)
            
            # Calcular cambio
            cambio = monto_pagado_float - total_orden
            print(f"DEBUG: Cálculo - Total: {total_orden}, Pagado: {monto_pagado_float}, Cambio: {cambio}")
            
            # Cliente ID
            cliente_id = None
            nombre_final = cliente_nombre
            
            if self.cliente_actual:
                cliente_id = self.cliente_actual['id']
                nombre_final = self.cliente_actual['nombre']
            
            print(f"DEBUG: cliente_id: {cliente_id}, nombre_final: {nombre_final}")
            
            # Crear orden
            query = """
            INSERT INTO ordenes (numero_orden, cliente_id, cliente_nombre, tipo_orden,
                               subtotal, total, metodo_pago, monto_pagado, cambio,
                               turno_id, cajero_id, fecha_orden)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """
            valores = (
                numero_orden, cliente_id, nombre_final, self.tipo_orden,
                total_orden, total_orden,
                metodo_pago, monto_pagado_float, cambio,
                auth.current_turno['id'], auth.current_user['id']
            )
            print(f"DEBUG: Valores para inserción: {valores}")
            
            orden_id = db.execute_one(query, valores)
            print(f"DEBUG: ID de orden creada: {orden_id}")
            
            if not orden_id:
                raise Exception("No se pudo crear la orden en la base de datos")
            
            # Crear detalles de la orden
            for i, item in enumerate(self.orden_actual['items']):
                print(f"DEBUG: Insertando item {i+1}: {item['nombre']}")
                
                # Convertir precios a float también
                precio_unitario = float(item['precio'])
                subtotal_item = float(item['subtotal'])
                
                query = """
                INSERT INTO orden_detalles (orden_id, producto_id, cantidad, precio_unitario, subtotal)
                VALUES (%s, %s, %s, %s, %s)
                """
                detalle_result = db.execute_one(query, (
                    orden_id, item['id'], item['cantidad'], precio_unitario, subtotal_item
                ))
                print(f"DEBUG: Detalle insertado con ID: {detalle_result}")
            
            # Imprimir tickets
            print("DEBUG: Generando tickets...")
            ticket_path = printer.print_customer_ticket(orden_id)
            comanda_path = printer.print_kitchen_ticket(orden_id)
            print(f"DEBUG: Tickets generados - Venta: {ticket_path}, Comanda: {comanda_path}")
            
            messagebox.showinfo("Éxito", f"Orden {numero_orden} creada exitosamente\nTickets generados correctamente")
            
            # Limpiar orden actual
            self.orden_actual = {'items': [], 'total': 0.0}
            self.cliente_actual = None
            
            # Volver al menú principal
            self.show_cajero_menu()
            
        except Exception as e:
            error_msg = f"Error creando la orden: {str(e)}"
            print(f"ERROR: {error_msg}")
            messagebox.showerror("Error", error_msg)

    def show_historial(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Historial de Órdenes del Día")
        dialog.geometry("900x650")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Título
        tk.Label(main_frame, text="📋 HISTORIAL DE ÓRDENES DEL DÍA", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        # Obtener órdenes del día
        query = """
        SELECT o.*, c.nombre as cliente_nombre, c.telefono
        FROM ordenes o
        LEFT JOIN clientes c ON o.cliente_id = c.id
        WHERE DATE(o.fecha_orden) = CURDATE() AND o.turno_id = %s
        ORDER BY o.fecha_orden DESC
        """
        ordenes = db.execute_query(query, (auth.current_turno['id'],))
        
        # Frame para treeview
        tree_frame = tk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True, pady=(0,15))
        
        # Crear treeview con scrollbar
        columns = ('numero_orden', 'cliente', 'tipo', 'total', 'pago', 'hora')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        tree.heading('numero_orden', text='# Orden')
        tree.heading('cliente', text='Cliente')
        tree.heading('tipo', text='Tipo')
        tree.heading('total', text='Total')
        tree.heading('pago', text='Método Pago')
        tree.heading('hora', text='Hora')
        
        tree.column('numero_orden', width=120)
        tree.column('cliente', width=180)
        tree.column('tipo', width=100)
        tree.column('total', width=100)
        tree.column('pago', width=100)
        tree.column('hora', width=80)
        
        # Scrollbar para treeview
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Llenar datos
        for orden in ordenes:
            tree.insert('', 'end', values=(
                orden['numero_orden'],
                orden['cliente_nombre'] or orden['cliente_nombre'] or 'Cliente Mostrador',
                orden['tipo_orden'].title(),
                f"${orden['total']:.2f}",
                orden['metodo_pago'].title(),
                orden['fecha_orden'].strftime('%H:%M')
            ), tags=(orden['id'],))
        
        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Frame para botones
        btn_frame = tk.Frame(main_frame, bg='#f8f9fa')
        btn_frame.pack(pady=(0,10))
        
        def get_selected_orden_id():
            selection = tree.selection()
            if selection:
                return tree.item(selection[0])['tags'][0]
            return None
        
        def preview_orden():
            orden_id = get_selected_orden_id()
            if not orden_id:
                messagebox.showwarning("Advertencia", "Por favor seleccione una orden")
                return
            self.show_orden_preview(orden_id)
        
        def reimprimir_venta():
            orden_id = get_selected_orden_id()
            if not orden_id:
                messagebox.showwarning("Advertencia", "Por favor seleccione una orden")
                return
            ticket_path = printer.print_customer_ticket(orden_id)
            if ticket_path:
                messagebox.showinfo("Éxito", f"Ticket reimpreso\nGuardado en: {ticket_path}")
        
        def reimprimir_comanda():
            orden_id = get_selected_orden_id()
            if not orden_id:
                messagebox.showwarning("Advertencia", "Por favor seleccione una orden")
                return
            comanda_path = printer.print_kitchen_ticket(orden_id)
            if comanda_path:
                messagebox.showinfo("Éxito", f"Comanda reimpresa\nGuardada en: {comanda_path}")
        
        # Botones organizados
        tk.Button(btn_frame, text="👁️ PREVISUALIZAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=15, height=1,
                 command=preview_orden).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="🎫 REIMPRIMIR VENTA", font=('Arial', 11, 'bold'),
                 bg='#2980b9', fg='white', width=18, height=1,
                 command=reimprimir_venta).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="📋 REIMPRIMIR COMANDA", font=('Arial', 11, 'bold'),
                 bg='#5dade2', fg='white', width=20, height=1,
                 command=reimprimir_comanda).pack(side='left', padx=5)
        
        def eliminar_orden():
            orden_id = get_selected_orden_id()
            if not orden_id:
                messagebox.showwarning("Advertencia", "Por favor seleccione una orden")
                return
            
            # Obtener información de la orden antes de eliminar
            query = "SELECT numero_orden, total FROM ordenes WHERE id = %s"
            orden_info = db.execute_one(query, (orden_id,))
            if not orden_info:
                messagebox.showerror("Error", "Orden no encontrada")
                return
            
            # Confirmar acción
            respuesta = messagebox.askyesno("Confirmar Eliminación", 
                f"⚠️ ATENCIÓN: Esta acción es IRREVERSIBLE\n\n"
                f"Orden: {orden_info['numero_orden']}\n"
                f"Total: ${float(orden_info['total']):.2f}\n\n"
                f"¿Está seguro de que desea eliminar esta orden?\n"
                f"Se eliminará completamente del sistema y no aparecerá en reportes.")
            
            if not respuesta:
                return
            
            # Pedir credenciales de administrador
            admin_user = self.admin_login_dialog("Eliminar Orden")
            if not admin_user:
                return
            
            # Proceder con la eliminación
            self.eliminar_orden_completa(orden_id, orden_info, dialog, tree)
        
        tk.Button(btn_frame, text="🗑️ ELIMINAR ORDEN", font=('Arial', 11, 'bold'),
                 bg='#e74c3c', fg='white', width=18, height=1,
                 command=eliminar_orden).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="CERRAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=dialog.destroy).pack(side='right', padx=5)
        
    def show_orden_preview(self, orden_id):
        try:
            # Obtener información completa de la orden
            query = """
            SELECT o.*, c.nombre as cliente_nombre, c.telefono, c.direccion,
                   u.nombre as cajero_nombre
            FROM ordenes o
            LEFT JOIN clientes c ON o.cliente_id = c.id
            LEFT JOIN usuarios u ON o.cajero_id = u.id
            WHERE o.id = %s
            """
            orden = db.execute_one(query, (orden_id,))
            
            if not orden:
                messagebox.showerror("Error", "Orden no encontrada")
                return
            
            # Obtener detalles de productos
            query = """
            SELECT od.*, p.nombre as producto_nombre
            FROM orden_detalles od
            JOIN productos p ON od.producto_id = p.id
            WHERE od.orden_id = %s
            ORDER BY od.id
            """
            detalles = db.execute_query(query, (orden_id,))
            
            # Crear ventana de previsualización
            preview_dialog = tk.Toplevel(self.root)
            preview_dialog.title(f"Previsualización - Orden {orden['numero_orden']}")
            preview_dialog.geometry("500x650")
            preview_dialog.resizable(False, False)
            preview_dialog.transient(self.root)
            preview_dialog.grab_set()
            
            # Configurar color de fondo
            preview_dialog.configure(bg='#f8f9fa')
            
            # Frame principal con scroll
            main_frame = tk.Frame(preview_dialog, bg='#f8f9fa')
            main_frame.pack(fill='both', expand=True, padx=15, pady=15)
            
            # Título
            tk.Label(main_frame, text="🎫 PREVISUALIZACIÓN DE ORDEN", font=('Arial', 14, 'bold'), 
                    bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
            
            # Frame para contenido con scroll
            canvas = tk.Canvas(main_frame, bg='white', relief='solid', bd=1)
            scrollbar = ttk.Scrollbar(main_frame, orient='vertical', command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='white')
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Contenido de la orden
            content_frame = tk.Frame(scrollable_frame, bg='white')
            content_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Información del encabezado
            header_frame = tk.LabelFrame(content_frame, text="Información General", font=('Arial', 11, 'bold'),
                                       bg='white', fg='#2c3e50')
            header_frame.pack(fill='x', pady=(0,15))
            
            info_frame = tk.Frame(header_frame, bg='white')
            info_frame.pack(fill='x', padx=10, pady=10)
            
            # Información en dos columnas
            # Columna izquierda
            left_col = tk.Frame(info_frame, bg='white')
            left_col.pack(side='left', fill='both', expand=True)
            
            tk.Label(left_col, text=f"📋 Orden: {orden['numero_orden']}", font=('Arial', 11, 'bold'), 
                    bg='white', fg='#3498db').pack(anchor='w', pady=2)
            tk.Label(left_col, text=f"📅 Fecha: {orden['fecha_orden'].strftime('%d/%m/%Y %H:%M')}", 
                    font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
            tk.Label(left_col, text=f"👤 Cajero: {orden['cajero_nombre']}", 
                    font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
            tk.Label(left_col, text=f"🏷️ Tipo: {orden['tipo_orden'].title()}", 
                    font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
            
            # Columna derecha
            right_col = tk.Frame(info_frame, bg='white')
            right_col.pack(side='right', fill='both', expand=True)
            
            if orden['cliente_nombre'] and orden['cliente_nombre'] != 'Cliente Mostrador':
                tk.Label(right_col, text=f"👥 Cliente: {orden['cliente_nombre']}", 
                        font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
                if orden['telefono']:
                    tk.Label(right_col, text=f"📱 Tel: {orden['telefono']}", 
                            font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
                if orden['direccion']:
                    tk.Label(right_col, text=f"📍 Dir: {orden['direccion']}", 
                            font=('Arial', 10), bg='white', wraplength=200).pack(anchor='w', pady=1)
            else:
                tk.Label(right_col, text="👥 Cliente: Mostrador", 
                        font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
            
            tk.Label(right_col, text=f"💳 Pago: {orden['metodo_pago'].title()}", 
                    font=('Arial', 10), bg='white').pack(anchor='w', pady=1)
            
            # Productos
            productos_frame = tk.LabelFrame(content_frame, text="Productos Ordenados", font=('Arial', 11, 'bold'),
                                          bg='white', fg='#2c3e50')
            productos_frame.pack(fill='x', pady=(0,15))
            
            # Headers de tabla
            table_header = tk.Frame(productos_frame, bg='#3498db')
            table_header.pack(fill='x', padx=10, pady=(10,0))
            
            tk.Label(table_header, text="Cant.", font=('Arial', 10, 'bold'), bg='#3498db', fg='white', width=6).pack(side='left')
            tk.Label(table_header, text="Producto", font=('Arial', 10, 'bold'), bg='#3498db', fg='white', width=25).pack(side='left')
            tk.Label(table_header, text="Precio", font=('Arial', 10, 'bold'), bg='#3498db', fg='white', width=10).pack(side='left')
            tk.Label(table_header, text="Subtotal", font=('Arial', 10, 'bold'), bg='#3498db', fg='white', width=10).pack(side='left')
            
            # Productos
            for i, detalle in enumerate(detalles):
                bg_color = '#f8f9fa' if i % 2 == 0 else 'white'
                product_row = tk.Frame(productos_frame, bg=bg_color)
                product_row.pack(fill='x', padx=10)
                
                tk.Label(product_row, text=str(detalle['cantidad']), font=('Arial', 10), 
                        bg=bg_color, width=6).pack(side='left', pady=2)
                tk.Label(product_row, text=detalle['producto_nombre'], font=('Arial', 10), 
                        bg=bg_color, width=25, anchor='w').pack(side='left', pady=2)
                tk.Label(product_row, text=f"${float(detalle['precio_unitario']):.2f}", font=('Arial', 10), 
                        bg=bg_color, width=10).pack(side='left', pady=2)
                tk.Label(product_row, text=f"${float(detalle['subtotal']):.2f}", font=('Arial', 10), 
                        bg=bg_color, width=10).pack(side='left', pady=2)
            
            # Total y pago
            totales_frame = tk.LabelFrame(content_frame, text="Totales y Pago", font=('Arial', 11, 'bold'),
                                        bg='white', fg='#2c3e50')
            totales_frame.pack(fill='x', pady=(0,15))
            
            totales_content = tk.Frame(totales_frame, bg='white')
            totales_content.pack(fill='x', padx=20, pady=15)
            
            # Total
            total_frame = tk.Frame(totales_content, bg='#e8f5e8', relief='solid', bd=1)
            total_frame.pack(fill='x', pady=2)
            tk.Label(total_frame, text=f"💰 TOTAL: ${float(orden['total']):.2f}", 
                    font=('Arial', 14, 'bold'), bg='#e8f5e8', fg='#27ae60').pack(pady=8)
            
            # Pago
            tk.Label(totales_content, text=f"Método de pago: {orden['metodo_pago'].title()}", 
                    font=('Arial', 11), bg='white').pack(anchor='w', pady=2)
            tk.Label(totales_content, text=f"Monto pagado: ${float(orden['monto_pagado']):.2f}", 
                    font=('Arial', 11), bg='white').pack(anchor='w', pady=2)
            
            if float(orden['cambio']) > 0:
                tk.Label(totales_content, text=f"Cambio entregado: ${float(orden['cambio']):.2f}", 
                        font=('Arial', 11), bg='white', fg='#3498db').pack(anchor='w', pady=2)
            
            # Empacar canvas y scrollbar
            canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            # Botón cerrar
            tk.Button(main_frame, text="CERRAR", font=('Arial', 11, 'bold'),
                     bg='#95a5a6', fg='white', width=15, height=1,
                     command=preview_dialog.destroy).pack(pady=(10,0))
            
            # Configurar scroll con rueda del mouse
            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            
            # Limpiar binding cuando se cierre la ventana
            def on_closing():
                canvas.unbind_all("<MouseWheel>")
                preview_dialog.destroy()
            
            preview_dialog.protocol("WM_DELETE_WINDOW", on_closing)
            
        except Exception as e:
            print(f"ERROR en show_orden_preview: {e}")
            messagebox.showerror("Error", f"Error mostrando previsualización: {str(e)}")

    def eliminar_orden_completa(self, orden_id, orden_info, parent_dialog, tree):
        try:
            print(f"DEBUG: Iniciando eliminación de orden {orden_id}")
            
            # Crear ventana de confirmación final - MÁS GRANDE
            confirm_dialog = tk.Toplevel(self.root)
            confirm_dialog.title("Confirmación Final - Eliminar Orden")
            confirm_dialog.geometry("550x650")
            confirm_dialog.resizable(False, False)
            confirm_dialog.transient(parent_dialog)
            confirm_dialog.grab_set()
            
            # Configurar color de fondo
            confirm_dialog.configure(bg='#f8f9fa')
            
            # Frame principal con scroll
            main_frame = tk.Frame(confirm_dialog, bg='#f8f9fa')
            main_frame.pack(fill='both', expand=True, padx=15, pady=15)
            
            # Título de advertencia - más compacto
            tk.Label(main_frame, text="⚠️ ELIMINACIÓN DE ORDEN", font=('Arial', 16, 'bold'), 
                    bg='#f8f9fa', fg='#e74c3c').pack(pady=(0,10))
            
            # Canvas y scrollbar para contenido scrolleable
            canvas = tk.Canvas(main_frame, bg='#f8f9fa', height=450)
            scrollbar = ttk.Scrollbar(main_frame, orient='vertical', command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Frame de advertencia - más compacto
            warning_frame = tk.LabelFrame(scrollable_frame, text="ADVERTENCIA IMPORTANTE", font=('Arial', 12, 'bold'),
                                        bg='#f8f9fa', fg='#e74c3c')
            warning_frame.pack(fill='x', pady=(0,10), padx=10)
            
            warning_text = """Esta acción eliminará COMPLETAMENTE la orden:

• Se eliminará la orden principal
• Se eliminarán todos los productos  
• No aparecerá en reportes ni historiales
• No se podrá recuperar
• Puede afectar cálculos de cortes

SOLO para corregir errores graves."""
            
            tk.Label(warning_frame, text=warning_text, font=('Arial', 10), 
                    bg='#f8f9fa', fg='#2c3e50', justify='left').pack(padx=10, pady=8)
            
            # Información de la orden - más compacto
            info_frame = tk.LabelFrame(scrollable_frame, text="Información de la Orden", font=('Arial', 12, 'bold'),
                                     bg='#f8f9fa', fg='#2c3e50')
            info_frame.pack(fill='x', pady=(0,10), padx=10)
            
            info_content = tk.Frame(info_frame, bg='#f8f9fa')
            info_content.pack(fill='x', padx=10, pady=8)
            
            tk.Label(info_content, text=f"📋 Orden: {orden_info['numero_orden']}", 
                    font=('Arial', 12, 'bold'), bg='#f8f9fa', fg='#3498db').pack(anchor='w', pady=1)
            tk.Label(info_content, text=f"💰 Total: ${float(orden_info['total']):.2f}", 
                    font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
            tk.Label(info_content, text=f"👤 Turno: #{auth.current_turno['id'] if auth.current_turno else 'N/A'}", 
                    font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
            
            # Obtener más información de la orden para mostrar
            try:
                query = """
                SELECT o.*, c.nombre as cliente_nombre, u.nombre as cajero_nombre,
                       COUNT(od.id) as total_productos
                FROM ordenes o
                LEFT JOIN clientes c ON o.cliente_id = c.id
                LEFT JOIN usuarios u ON o.cajero_id = u.id
                LEFT JOIN orden_detalles od ON o.id = od.orden_id
                WHERE o.id = %s
                GROUP BY o.id
                """
                orden_detalle = db.execute_one(query, (orden_id,))
                
                if orden_detalle:
                    tk.Label(info_content, text=f"📅 Fecha: {orden_detalle['fecha_orden'].strftime('%d/%m/%Y %H:%M')}", 
                            font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
                    tk.Label(info_content, text=f"🏷️ Tipo: {orden_detalle['tipo_orden'].title()}", 
                            font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
                    tk.Label(info_content, text=f"👤 Cajero: {orden_detalle['cajero_nombre']}", 
                            font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
                    if orden_detalle['cliente_nombre']:
                        tk.Label(info_content, text=f"👥 Cliente: {orden_detalle['cliente_nombre']}", 
                                font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
                    tk.Label(info_content, text=f"🛒 Productos: {orden_detalle['total_productos']}", 
                            font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
                    tk.Label(info_content, text=f"💳 Pago: {orden_detalle['metodo_pago'].title()}", 
                            font=('Arial', 11), bg='#f8f9fa').pack(anchor='w', pady=1)
            except:
                pass  # Si hay error obteniendo detalles, continúa con información básica
            
            # Razón de eliminación
            reason_frame = tk.LabelFrame(scrollable_frame, text="Razón de Eliminación (Obligatorio)", font=('Arial', 12, 'bold'),
                                       bg='#f8f9fa', fg='#2c3e50')
            reason_frame.pack(fill='x', pady=(0,10), padx=10)
            
            tk.Label(reason_frame, text="Explique detalladamente por qué elimina esta orden:", 
                    font=('Arial', 10), bg='#f8f9fa', fg='#7f8c8d').pack(padx=10, pady=(5,0))
            
            reason_entry = tk.Text(reason_frame, height=4, font=('Arial', 10), relief='solid', bd=1)
            reason_entry.pack(fill='x', padx=10, pady=(5,10))
            reason_entry.insert('1.0', 'Motivo: ')
            
            # Confirmación final
            final_frame = tk.LabelFrame(scrollable_frame, text="Confirmación Final", font=('Arial', 12, 'bold'),
                                      bg='#f8f9fa', fg='#e74c3c')
            final_frame.pack(fill='x', pady=(0,10), padx=10)
            
            confirm_text = """Al presionar "ELIMINAR DEFINITIVAMENTE":

✓ Acepto que esta acción es IRREVERSIBLE
✓ Entiendo que puede afectar reportes y cortes
✓ Confirmo que es necesario eliminar esta orden
✓ Asumo la responsabilidad de esta acción"""
            
            tk.Label(final_frame, text=confirm_text, font=('Arial', 10), 
                    bg='#f8f9fa', fg='#2c3e50', justify='left').pack(padx=10, pady=8)
            
            # Empacar canvas y scrollbar
            canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            def ejecutar_eliminacion():
                reason = reason_entry.get('1.0', tk.END).strip()
                if len(reason) < 15 or reason == 'Motivo:' or 'Motivo:' in reason and len(reason) < 20:
                    messagebox.showerror("Error", "Por favor proporcione una razón detallada para la eliminación\n(mínimo 15 caracteres útiles)")
                    reason_entry.focus()
                    return
                
                try:
                    print(f"DEBUG: Eliminando orden {orden_id} - Razón: {reason}")
                    
                    # Iniciar transacción para eliminar todo
                    # Primero eliminar detalles de la orden
                    query = "DELETE FROM orden_detalles WHERE orden_id = %s"
                    result_detalles = db.execute_query(query, (orden_id,))
                    print(f"DEBUG: Detalles eliminados: {result_detalles}")
                    
                    # Luego eliminar la orden principal
                    query = "DELETE FROM ordenes WHERE id = %s"
                    result_orden = db.execute_query(query, (orden_id,))
                    print(f"DEBUG: Orden eliminada: {result_orden}")
                    
                    if result_orden:
                        # Registrar la eliminación en un log
                        print(f"AUDIT: Orden {orden_info['numero_orden']} eliminada por admin {auth.current_user['nombre']} - Razón: {reason}")
                        
                        messagebox.showinfo("Éxito", 
                            f"✅ Orden {orden_info['numero_orden']} eliminada exitosamente\n\n"
                            f"La orden ya no aparecerá en reportes ni historiales.\n"
                            f"Acción registrada en logs del sistema.")
                        
                        # Cerrar diálogos
                        confirm_dialog.destroy()
                        
                        # Actualizar el historial eliminando la fila
                        selection = tree.selection()
                        if selection:
                            tree.delete(selection[0])
                    else:
                        messagebox.showerror("Error", "No se pudo eliminar la orden de la base de datos")
                        
                except Exception as e:
                    print(f"ERROR eliminando orden: {e}")
                    messagebox.showerror("Error", f"Error eliminando orden: {str(e)}")
            
            def cancelar_eliminacion():
                confirm_dialog.destroy()
            
            # Botones fijos en la parte inferior
            button_frame = tk.Frame(main_frame, bg='#f8f9fa')
            button_frame.pack(fill='x', pady=(10,0))
            
            tk.Button(button_frame, text="CANCELAR", font=('Arial', 12, 'bold'),
                     bg='#95a5a6', fg='white', width=15, height=1,
                     command=cancelar_eliminacion).pack(side='left', padx=5)
            
            tk.Button(button_frame, text="⚠️ ELIMINAR DEFINITIVAMENTE", font=('Arial', 12, 'bold'),
                     bg='#e74c3c', fg='white', width=25, height=1,
                     command=ejecutar_eliminacion).pack(side='right', padx=5)
            
            # Configurar scroll con rueda del mouse
            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            
            # Limpiar binding cuando se cierre la ventana
            def on_closing():
                canvas.unbind_all("<MouseWheel>")
                confirm_dialog.destroy()
            
            confirm_dialog.protocol("WM_DELETE_WINDOW", on_closing)
            
            reason_entry.focus()
            reason_entry.mark_set(tk.INSERT, '1.8')  # Posicionar cursor después de "Motivo: "
            
        except Exception as e:
            print(f"ERROR en eliminar_orden_completa: {e}")
            messagebox.showerror("Error", f"Error preparando eliminación: {str(e)}")

    def show_arqueo(self):
        # Pedir credenciales de administrador
        admin_dialog = self.admin_login_dialog("Arqueo de Caja")
        if not admin_dialog:
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Arqueo de Caja")
        dialog.geometry("380x280")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal compacto
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="💰 ARQUEO DE CAJA", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,8))
        
        tk.Label(main_frame, text="Retirar efectivo durante el turno", font=('Arial', 10), 
                bg='#f8f9fa', fg='#7f8c8d').pack(pady=(0,15))
        
        # Frame para monto - más compacto
        monto_frame = tk.LabelFrame(main_frame, text="Monto a Retirar", font=('Arial', 11),
                                   bg='#f8f9fa', fg='#2c3e50')
        monto_frame.pack(pady=(0,15), fill='x')
        
        entry_frame = tk.Frame(monto_frame, bg='#f8f9fa')
        entry_frame.pack(pady=10)
        
        tk.Label(entry_frame, text="$", font=('Arial', 16, 'bold'), bg='#f8f9fa').pack(side='left')
        amount_entry = tk.Entry(entry_frame, font=('Arial', 16), width=8, justify='center',
                               relief='solid', bd=1)
        amount_entry.pack(side='left', padx=5)
        
        def procesar_arqueo():
            try:
                amount_text = amount_entry.get().strip()
                if not amount_text:
                    messagebox.showerror("Error", "Por favor ingrese el monto a retirar")
                    amount_entry.focus()
                    return
                
                monto = float(amount_text)
                if monto <= 0:
                    messagebox.showerror("Error", "El monto debe ser mayor a 0")
                    amount_entry.focus()
                    return
                
                query = """
                INSERT INTO arqueos (turno_id, monto, administrador_id)
                VALUES (%s, %s, %s)
                """
                result = db.execute_one(query, (
                    auth.current_turno['id'], monto, admin_dialog['id']
                ))
                
                if result:
                    messagebox.showinfo("Éxito", f"Arqueo registrado exitosamente\nMonto retirado: ${monto:.2f}")
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "No se pudo registrar el arqueo")
            except ValueError:
                messagebox.showerror("Error", "Por favor ingrese un monto válido")
                amount_entry.focus()
        
        # Botones - más compactos
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(pady=(0,5), fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="PROCESAR ARQUEO", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=15, height=1,
                 command=procesar_arqueo).pack(side='right', padx=5)
        
        amount_entry.focus()
        amount_entry.bind('<Return>', lambda e: procesar_arqueo())

    def show_corte_caja(self):
        # Pedir credenciales de administrador
        admin_user = self.admin_login_dialog("Corte de Caja")
        if not admin_user:
            return
        
        self.procesar_corte_caja(admin_user)

    def procesar_corte_caja(self, admin_user):
        dialog = tk.Toplevel(self.root)
        dialog.title("Corte de Caja")
        dialog.geometry("420x350")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="🏦 CORTE DE CAJA", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        # Paso 1: Cajero ingresa cantidades
        step1_frame = tk.LabelFrame(main_frame, text="Paso 1: Cantidades Finales", font=('Arial', 11),
                                   bg='#f8f9fa', fg='#2c3e50')
        step1_frame.pack(fill='x', pady=(0,15))
        
        tk.Label(step1_frame, text="Ingrese las cantidades que tiene al final del turno:", 
                font=('Arial', 10), bg='#f8f9fa', fg='#7f8c8d').pack(pady=(10,15))
        
        # Efectivo
        efectivo_frame = tk.Frame(step1_frame, bg='#f8f9fa')
        efectivo_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Label(efectivo_frame, text="💵 Efectivo en caja:", font=('Arial', 11), 
                bg='#f8f9fa').pack(side='left')
        tk.Label(efectivo_frame, text="$", font=('Arial', 12, 'bold'), 
                bg='#f8f9fa').pack(side='right', padx=(5,0))
        efectivo_entry = tk.Entry(efectivo_frame, font=('Arial', 11), width=10, 
                                 relief='solid', bd=1, justify='right')
        efectivo_entry.pack(side='right')
        
        # Tarjeta
        tarjeta_frame = tk.Frame(step1_frame, bg='#f8f9fa')
        tarjeta_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Label(tarjeta_frame, text="💳 Total tarjetas:", font=('Arial', 11), 
                bg='#f8f9fa').pack(side='left')
        tk.Label(tarjeta_frame, text="$", font=('Arial', 12, 'bold'), 
                bg='#f8f9fa').pack(side='right', padx=(5,0))
        tarjeta_entry = tk.Entry(tarjeta_frame, font=('Arial', 11), width=10, 
                                relief='solid', bd=1, justify='right')
        tarjeta_entry.pack(side='right')
        
        # Advertencia
        warning_frame = tk.Frame(step1_frame, bg='#fff2cd', relief='solid', bd=1)
        warning_frame.pack(fill='x', padx=10, pady=(15,10))
        
        warning_label = tk.Label(warning_frame, 
                               text="⚠️ PRECAUCIÓN: Cualquier faltante será cobrado",
                               font=('Arial', 10, 'bold'), bg='#fff2cd', fg='#856404')
        warning_label.pack(pady=8)
        
        def paso2():
            try:
                efectivo_text = efectivo_entry.get().strip()
                tarjeta_text = tarjeta_entry.get().strip()
                
                if not efectivo_text:
                    messagebox.showerror("Error", "Por favor ingrese el monto de efectivo")
                    efectivo_entry.focus()
                    return
                
                if not tarjeta_text:
                    messagebox.showerror("Error", "Por favor ingrese el monto de tarjetas")
                    tarjeta_entry.focus()
                    return
                
                efectivo = float(efectivo_text)
                tarjeta = float(tarjeta_text)
                
                if efectivo < 0:
                    messagebox.showerror("Error", "El monto de efectivo no puede ser negativo")
                    efectivo_entry.focus()
                    return
                
                if tarjeta < 0:
                    messagebox.showerror("Error", "El monto de tarjetas no puede ser negativo")
                    tarjeta_entry.focus()
                    return
                
                print(f"DEBUG: Procediendo a paso 2 con efectivo: {efectivo}, tarjeta: {tarjeta}")
                self.show_corte_paso2(dialog, admin_user, efectivo, tarjeta)
                
            except ValueError:
                messagebox.showerror("Error", "Por favor ingrese montos válidos (solo números)")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x', pady=(0,5))
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="CONTINUAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=paso2).pack(side='right', padx=5)
        
        efectivo_entry.focus()
        efectivo_entry.bind('<Return>', lambda e: tarjeta_entry.focus())
        tarjeta_entry.bind('<Return>', lambda e: paso2())

    def show_corte_paso2(self, parent_dialog, admin_user, efectivo_cajero, tarjeta_cajero):
        parent_dialog.destroy()
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Corte de Caja - Confirmación")
        dialog.geometry("520x650")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="💰 CONFIRMACIÓN DE CORTE", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        try:
            print("DEBUG: Iniciando cálculo de corte de caja")
            print(f"DEBUG: Turno ID: {auth.current_turno['id'] if auth.current_turno else 'None'}")
            
            # Calcular totales esperados - convertir todo a float
            # Obtener total de ventas
            query = """
            SELECT 
                COALESCE(SUM(CASE WHEN metodo_pago = 'efectivo' THEN total ELSE 0 END), 0) as ventas_efectivo,
                COALESCE(SUM(CASE WHEN metodo_pago = 'tarjeta' THEN total ELSE 0 END), 0) as ventas_tarjeta,
                COALESCE(SUM(total), 0) as total_ventas
            FROM ordenes 
            WHERE turno_id = %s
            """
            ventas = db.execute_one(query, (auth.current_turno['id'],))
            print(f"DEBUG: Ventas obtenidas: {ventas}")
            
            # Convertir a float para evitar problemas de tipo
            ventas_efectivo = float(ventas['ventas_efectivo']) if ventas else 0.0
            ventas_tarjeta = float(ventas['ventas_tarjeta']) if ventas else 0.0
            total_ventas = float(ventas['total_ventas']) if ventas else 0.0
            
            # Obtener arqueos
            query = "SELECT COALESCE(SUM(monto), 0) as total_arqueos FROM arqueos WHERE turno_id = %s"
            arqueos_result = db.execute_one(query, (auth.current_turno['id'],))
            total_arqueos = float(arqueos_result['total_arqueos']) if arqueos_result else 0.0
            
            print(f"DEBUG: Arqueos: {total_arqueos}")
            
            # Obtener fondo inicial y convertir a float
            fondo_inicial = float(auth.current_turno['fondo_inicial'])
            print(f"DEBUG: Fondo inicial: {fondo_inicial}")
            
            # Calcular esperados
            efectivo_esperado = fondo_inicial + ventas_efectivo - total_arqueos
            tarjeta_esperada = ventas_tarjeta
            total_esperado = efectivo_esperado + tarjeta_esperada
            total_real = efectivo_cajero + tarjeta_cajero
            diferencia = total_esperado - total_real
            
            print(f"DEBUG: Cálculos - Efectivo esperado: {efectivo_esperado}, Tarjeta esperada: {tarjeta_esperada}")
            print(f"DEBUG: Total esperado: {total_esperado}, Total real: {total_real}, Diferencia: {diferencia}")
            
            # Mostrar resumen
            info_frame = tk.LabelFrame(main_frame, text="Resumen del Turno", font=('Arial', 12), 
                                     bg='#f8f9fa', fg='#2c3e50')
            info_frame.pack(fill='both', expand=True, pady=(0,15))
            
            # Crear texto del resumen
            info_text = f"""INFORMACIÓN DEL TURNO:
Turno #{auth.current_turno['id']}
Fondo inicial: ${fondo_inicial:.2f}
Ventas efectivo: ${ventas_efectivo:.2f}
Ventas tarjeta: ${ventas_tarjeta:.2f}
Total arqueos: ${total_arqueos:.2f}

CANTIDADES ESPERADAS:
Efectivo esperado: ${efectivo_esperado:.2f}
Tarjeta esperada: ${tarjeta_esperada:.2f}
Total esperado: ${total_esperado:.2f}

CANTIDADES REPORTADAS:
Efectivo reportado: ${efectivo_cajero:.2f}
Tarjeta reportada: ${tarjeta_cajero:.2f}
Total reportado: ${total_real:.2f}

DIFERENCIA: ${diferencia:.2f}"""

            if diferencia > 0.01:
                info_text += f"\n\n⚠️ FALTANTE DE: ${diferencia:.2f}"
            elif diferencia < -0.01:
                info_text += f"\n\n✅ SOBRANTE DE: ${abs(diferencia):.2f}"
            else:
                info_text += f"\n\n✅ CORTE CUADRADO"
            
            text_widget = tk.Text(info_frame, height=18, width=55, font=('Arial', 10),
                                 wrap=tk.WORD, relief='solid', bd=1)
            text_widget.insert('1.0', info_text)
            text_widget.config(state='disabled')
            text_widget.pack(padx=10, pady=10, fill='both', expand=True)
            
            # Pedir confirmación del administrador
            admin_frame = tk.LabelFrame(main_frame, text="Confirmación de Administrador", font=('Arial', 11),
                                       bg='#f8f9fa', fg='#2c3e50')
            admin_frame.pack(fill='x', pady=(0,10))
            
            tk.Label(admin_frame, text="Contraseña de administrador:", font=('Arial', 10), 
                    bg='#f8f9fa').pack(anchor='w', padx=10, pady=(10,2))
            
            admin_pass_entry = tk.Entry(admin_frame, show='*', font=('Arial', 11), relief='solid', bd=1)
            admin_pass_entry.pack(fill='x', padx=10, pady=(0,10))
            
            def finalizar_corte():
                try:
                    password = admin_pass_entry.get().strip()
                    if not password:
                        messagebox.showerror("Error", "Por favor ingrese la contraseña")
                        admin_pass_entry.focus()
                        return
                    
                    # Verificar contraseña del administrador nuevamente
                    query = "SELECT * FROM usuarios WHERE id = %s AND password = %s"
                    admin_check = db.execute_one(query, (admin_user['id'], password))
                    
                    if not admin_check:
                        messagebox.showerror("Error", "Contraseña de administrador incorrecta")
                        admin_pass_entry.focus()
                        return
                    
                    print("DEBUG: Contraseña verificada, procediendo a cerrar turno...")
                    print(f"DEBUG: Valores - Efectivo: {efectivo_cajero}, Tarjeta: {tarjeta_cajero}")
                    
                    # Guardar el ID del turno antes de cerrarlo
                    turno_id_para_corte = auth.current_turno['id'] if auth.current_turno else None
                    
                    # Cerrar turno
                    resultado_cerrar = auth.cerrar_turno(efectivo_cajero, tarjeta_cajero)
                    print(f"DEBUG: Resultado cerrar_turno: {resultado_cerrar}")
                    
                    if resultado_cerrar:
                        # Imprimir corte
                        print(f"DEBUG: Imprimiendo corte para turno: {turno_id_para_corte}")
                        
                        if turno_id_para_corte:
                            corte_path = printer.print_corte_caja(turno_id_para_corte)
                            print(f"DEBUG: Corte impreso en: {corte_path}")
                        
                        messagebox.showinfo("Éxito", "Corte de caja realizado correctamente\nSesión cerrada automáticamente")
                        dialog.destroy()
                        
                        # Hacer logout después del corte
                        auth.logout()
                        self.show_login()
                    else:
                        messagebox.showerror("Error", "No se pudo procesar el corte de caja.\nVerifique la conexión a la base de datos.")
                        
                except Exception as e:
                    print(f"ERROR en finalizar_corte: {e}")
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Error", f"Error finalizando corte: {str(e)}")
            
            # Botones
            button_frame = tk.Frame(main_frame, bg='#f8f9fa')
            button_frame.pack(fill='x', pady=(0,5))
            
            tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                     bg='#95a5a6', fg='white', width=12, height=1,
                     command=dialog.destroy).pack(side='left', padx=5)
            
            tk.Button(button_frame, text="FINALIZAR CORTE", font=('Arial', 11, 'bold'),
                     bg='#3498db', fg='white', width=15, height=1,
                     command=finalizar_corte).pack(side='right', padx=5)
            
            admin_pass_entry.focus()
            admin_pass_entry.bind('<Return>', lambda e: finalizar_corte())
            
        except Exception as e:
            print(f"ERROR en show_corte_paso2: {e}")
            error_frame = tk.Frame(main_frame, bg='#f8f9fa')
            error_frame.pack(fill='both', expand=True)
            
            tk.Label(error_frame, text="❌ ERROR CALCULANDO CORTE", font=('Arial', 14, 'bold'),
                    bg='#f8f9fa', fg='#e74c3c').pack(pady=20)
            
            tk.Label(error_frame, text=f"Error: {str(e)}", font=('Arial', 10),
                    bg='#f8f9fa', fg='#7f8c8d', wraplength=400).pack(pady=10)
            
            tk.Button(error_frame, text="CERRAR", font=('Arial', 11, 'bold'),
                     bg='#95a5a6', fg='white', command=dialog.destroy).pack(pady=20)

    def admin_login_dialog(self, title):
        dialog = tk.Toplevel(self.root)
        dialog.title(f"{title} - Login Administrador")
        dialog.geometry("350x250")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar color de fondo
        dialog.configure(bg='#f8f9fa')
        
        result = None
        
        # Frame principal compacto
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="🔐 CREDENCIALES DE ADMINISTRADOR", font=('Arial', 12, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,15))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(pady=(0,15))
        
        tk.Label(fields_frame, text="Usuario:", font=('Arial', 11), 
                bg='#f8f9fa', fg='#2c3e50').grid(row=0, column=0, sticky='w', pady=6)
        user_entry = tk.Entry(fields_frame, font=('Arial', 11), width=15, relief='solid', bd=1)
        user_entry.grid(row=0, column=1, pady=6, padx=8)
        
        tk.Label(fields_frame, text="Contraseña:", font=('Arial', 11), 
                bg='#f8f9fa', fg='#2c3e50').grid(row=1, column=0, sticky='w', pady=6)
        pass_entry = tk.Entry(fields_frame, show='*', font=('Arial', 11), width=15, relief='solid', bd=1)
        pass_entry.grid(row=1, column=1, pady=6, padx=8)
        
        def verificar():
            nonlocal result
            username = user_entry.get()
            password = pass_entry.get()
            
            query = """
            SELECT * FROM usuarios 
            WHERE username = %s AND password = %s AND tipo IN ('administrador', 'superusuario')
            """
            user = db.execute_one(query, (username, password))
            
            if user:
                result = user
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Credenciales incorrectas")
        
        # Botones - más compactos
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(pady=(0,5), fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=10, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="VERIFICAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=verificar).pack(side='right', padx=5)
        
        user_entry.focus()
        user_entry.bind('<Return>', lambda e: pass_entry.focus())
        pass_entry.bind('<Return>', lambda e: verificar())
        
        dialog.wait_window()
        return result

    def show_admin_menu(self):
        self.clear_window()
        
        # Header
        header_frame = tk.Frame(self.root, bg='#34495e', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        user_info = f"Administrador: {auth.current_user['nombre']}"
        tk.Label(header_frame, text=user_info, font=('Arial', 12, 'bold'), 
                bg='#34495e', fg='white').pack(side='left', padx=20, pady=20)
        
        tk.Button(header_frame, text="Cerrar Sesión", command=self.logout).pack(side='right', padx=20, pady=15)
        
        # Menú principal
        main_frame = tk.Frame(self.root, bg='#f5f5f5')
        main_frame.pack(expand=True, fill='both', padx=20, pady=20)
        
        tk.Label(main_frame, text="Panel de Administración", 
                font=('Arial', 18, 'bold'), bg='#f5f5f5').pack(pady=20)
        
        # Grid de botones
        buttons_frame = tk.Frame(main_frame, bg='#f5f5f5')
        buttons_frame.pack(expand=True)
        
        # Fila 1
        row1 = tk.Frame(buttons_frame, bg='#f5f5f5')
        row1.pack(pady=10)
        
        tk.Button(row1, text="Administrar Menú", font=('Arial', 12, 'bold'),
                 bg='#3498db', fg='white', width=20, height=3,
                 command=self.admin_menu).pack(side='left', padx=10)
        
        tk.Button(row1, text="Administrar Usuarios", font=('Arial', 12, 'bold'),
                 bg='#2980b9', fg='white', width=20, height=3,
                 command=self.admin_usuarios).pack(side='left', padx=10)
        
        # Fila 2
        row2 = tk.Frame(buttons_frame, bg='#f5f5f5')
        row2.pack(pady=10)
        
        tk.Button(row2, text="Administrar Ventas", font=('Arial', 12, 'bold'),
                 bg='#5dade2', fg='white', width=20, height=3,
                 command=self.admin_ventas).pack(side='left', padx=10)
        
        tk.Button(row2, text="Administrar Clientes", font=('Arial', 12, 'bold'),
                 bg='#85c1e9', fg='white', width=20, height=3,
                 command=self.admin_clientes).pack(side='left', padx=10)

    def admin_menu(self):
        self.show_menu_administration()
    
    def show_menu_administration(self):
        self.clear_window()
        
        # Header
        header_frame = tk.Frame(self.root, bg='#34495e', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        user_info = f"Admin - Gestión de Menú: {auth.current_user['nombre']}"
        tk.Label(header_frame, text=user_info, font=('Arial', 12, 'bold'), 
                bg='#34495e', fg='white').pack(side='left', padx=20, pady=20)
        
        # Botón volver
        tk.Button(header_frame, text="← Volver", font=('Arial', 10), 
                 command=self.show_admin_menu).pack(side='right', padx=20, pady=15)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Título
        tk.Label(main_frame, text="🍽️ ADMINISTRACIÓN DE MENÚ", font=('Arial', 18, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Frame para pestañas/secciones
        notebook_frame = tk.Frame(main_frame, bg='#f8f9fa')
        notebook_frame.pack(fill='both', expand=True)
        
        # Botones de sección
        section_frame = tk.Frame(notebook_frame, bg='#f8f9fa')
        section_frame.pack(fill='x', pady=(0,10))
        
        self.current_section = tk.StringVar(value='categorias')
        
        tk.Radiobutton(section_frame, text="📁 CATEGORÍAS", variable=self.current_section, 
                      value='categorias', font=('Arial', 12, 'bold'), bg='#f8f9fa',
                      command=self.refresh_menu_content).pack(side='left', padx=20)
        tk.Radiobutton(section_frame, text="🍕 PRODUCTOS", variable=self.current_section, 
                      value='productos', font=('Arial', 12, 'bold'), bg='#f8f9fa',
                      command=self.refresh_menu_content).pack(side='left', padx=20)
        tk.Radiobutton(section_frame, text="📊 REPORTES", variable=self.current_section, 
                      value='reportes', font=('Arial', 12, 'bold'), bg='#f8f9fa',
                      command=self.refresh_menu_content).pack(side='left', padx=20)
        
        # Contenedor para el contenido dinámico
        self.menu_content_frame = tk.Frame(notebook_frame, bg='#f8f9fa')
        self.menu_content_frame.pack(fill='both', expand=True)
        
        # Cargar contenido inicial
        self.refresh_menu_content()
    
    def refresh_menu_content(self):
        # Limpiar contenido anterior
        for widget in self.menu_content_frame.winfo_children():
            widget.destroy()
        
        section = self.current_section.get()
        
        if section == 'categorias':
            self.show_categorias_management()
        elif section == 'productos':
            self.show_productos_management()
        elif section == 'reportes':
            self.show_menu_reports()
    
    def show_categorias_management(self):
        # Frame principal para categorías
        frame = tk.Frame(self.menu_content_frame, bg='#f8f9fa')
        frame.pack(fill='both', expand=True)
        
        # Botones de acción
        action_frame = tk.Frame(frame, bg='#f8f9fa')
        action_frame.pack(fill='x', pady=(0,15))
        
        tk.Button(action_frame, text="➕ NUEVA CATEGORÍA", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=18, height=1,
                 command=self.nueva_categoria).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="✏️ EDITAR", font=('Arial', 11, 'bold'),
                 bg='#2980b9', fg='white', width=12, height=1,
                 command=self.editar_categoria).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="🗑️ ELIMINAR", font=('Arial', 11, 'bold'),
                 bg='#e74c3c', fg='white', width=12, height=1,
                 command=self.eliminar_categoria).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="📄 EXPORTAR", font=('Arial', 11, 'bold'),
                 bg='#27ae60', fg='white', width=12, height=1,
                 command=self.exportar_categorias).pack(side='right', padx=5)
        
        # Lista de categorías
        list_frame = tk.Frame(frame, bg='#f8f9fa')
        list_frame.pack(fill='both', expand=True)
        
        # Treeview para categorías (sin descripción)
        columns = ('id', 'nombre', 'activo', 'productos')
        self.categorias_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        self.categorias_tree.heading('id', text='ID')
        self.categorias_tree.heading('nombre', text='Nombre')
        self.categorias_tree.heading('activo', text='Estado')
        self.categorias_tree.heading('productos', text='Productos')
        
        self.categorias_tree.column('id', width=50)
        self.categorias_tree.column('nombre', width=200)
        self.categorias_tree.column('activo', width=100)
        self.categorias_tree.column('productos', width=100)
        
        scrollbar_cat = ttk.Scrollbar(list_frame, orient='vertical', command=self.categorias_tree.yview)
        self.categorias_tree.configure(yscrollcommand=scrollbar_cat.set)
        
        self.categorias_tree.pack(side='left', fill='both', expand=True)
        scrollbar_cat.pack(side='right', fill='y')
        
        # Cargar datos
        self.load_categorias()
    
    def show_productos_management(self):
        frame = tk.Frame(self.menu_content_frame, bg='#f8f9fa')
        frame.pack(fill='both', expand=True)
        
        # Filtros
        filter_frame = tk.Frame(frame, bg='#f8f9fa')
        filter_frame.pack(fill='x', pady=(0,10))
        
        tk.Label(filter_frame, text="🔍 Filtrar por categoría:", font=('Arial', 11), 
                bg='#f8f9fa').pack(side='left', padx=5)
        
        self.categoria_filter = ttk.Combobox(filter_frame, width=20, state='readonly')
        self.categoria_filter.pack(side='left', padx=5)
        self.categoria_filter.bind('<<ComboboxSelected>>', lambda e: self.load_productos())
        
        # Botones de acción
        action_frame = tk.Frame(frame, bg='#f8f9fa')
        action_frame.pack(fill='x', pady=(0,15))
        
        tk.Button(action_frame, text="➕ NUEVO PRODUCTO", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=18, height=1,
                 command=self.nuevo_producto).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="✏️ EDITAR", font=('Arial', 11, 'bold'),
                 bg='#2980b9', fg='white', width=12, height=1,
                 command=self.editar_producto).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="🗑️ ELIMINAR", font=('Arial', 11, 'bold'),
                 bg='#e74c3c', fg='white', width=12, height=1,
                 command=self.eliminar_producto).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="📄 EXPORTAR", font=('Arial', 11, 'bold'),
                 bg='#27ae60', fg='white', width=12, height=1,
                 command=self.exportar_productos).pack(side='right', padx=5)
        
        tk.Button(action_frame, text="📥 IMPORTAR", font=('Arial', 11, 'bold'),
                 bg='#f39c12', fg='white', width=12, height=1,
                 command=self.importar_productos).pack(side='right', padx=5)
        
        # Lista de productos
        list_frame = tk.Frame(frame, bg='#f8f9fa')
        list_frame.pack(fill='both', expand=True)
        
        columns = ('id', 'nombre', 'categoria', 'precio', 'activo')
        self.productos_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        self.productos_tree.heading('id', text='ID')
        self.productos_tree.heading('nombre', text='Nombre')
        self.productos_tree.heading('categoria', text='Categoría')
        self.productos_tree.heading('precio', text='Precio')
        self.productos_tree.heading('activo', text='Estado')
        
        self.productos_tree.column('id', width=60)
        self.productos_tree.column('nombre', width=220)
        self.productos_tree.column('categoria', width=150)
        self.productos_tree.column('precio', width=100)
        self.productos_tree.column('activo', width=100)
        
        scrollbar_prod = ttk.Scrollbar(list_frame, orient='vertical', command=self.productos_tree.yview)
        self.productos_tree.configure(yscrollcommand=scrollbar_prod.set)
        
        self.productos_tree.pack(side='left', fill='both', expand=True)
        scrollbar_prod.pack(side='right', fill='y')
        
        # Cargar datos
        self.load_categoria_filter()
        self.load_productos()
    
    def show_menu_reports(self):
        frame = tk.Frame(self.menu_content_frame, bg='#f8f9fa')
        frame.pack(fill='both', expand=True)
        
        # Título de reportes
        tk.Label(frame, text="📊 REPORTES DE MENÚ", font=('Arial', 16, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Grid de botones de reportes
        reports_grid = tk.Frame(frame, bg='#f8f9fa')
        reports_grid.pack(expand=True)
        
        # Fila 1
        row1 = tk.Frame(reports_grid, bg='#f8f9fa')
        row1.pack(pady=10)
        
        tk.Button(row1, text="📈 PRODUCTOS MÁS VENDIDOS", font=('Arial', 12, 'bold'),
                 bg='#3498db', fg='white', width=25, height=3,
                 command=self.reporte_productos_vendidos).pack(side='left', padx=10)
        
        tk.Button(row1, text="💰 ANÁLISIS DE PRECIOS", font=('Arial', 12, 'bold'),
                 bg='#2980b9', fg='white', width=25, height=3,
                 command=self.reporte_analisis_precios).pack(side='left', padx=10)
        
        # Fila 2
        row2 = tk.Frame(reports_grid, bg='#f8f9fa')
        row2.pack(pady=10)
        
        tk.Button(row2, text="📋 INVENTARIO COMPLETO", font=('Arial', 12, 'bold'),
                 bg='#5dade2', fg='white', width=25, height=3,
                 command=self.reporte_inventario_completo).pack(side='left', padx=10)
        
    # ================== FUNCIONES DE CATEGORÍAS ==================
    
    def load_categorias(self):
        """Cargar categorías en el treeview"""
        try:
            # Limpiar tree
            for item in self.categorias_tree.get_children():
                self.categorias_tree.delete(item)
            
            # Obtener categorías con conteo de productos (sin descripción)
            query = """
            SELECT c.id, c.nombre, c.activo, COUNT(p.id) as total_productos
            FROM categorias c
            LEFT JOIN productos p ON c.id = p.categoria_id AND p.activo = 1
            GROUP BY c.id, c.nombre, c.activo
            ORDER BY c.nombre
            """
            categorias = db.execute_query(query)
            
            for categoria in categorias:
                estado = "✅ Activo" if categoria['activo'] else "❌ Inactivo"
                self.categorias_tree.insert('', 'end', values=(
                    categoria['id'],
                    categoria['nombre'],
                    estado,
                    categoria['total_productos']
                ))
                
        except Exception as e:
            messagebox.showerror("Error", f"Error cargando categorías: {str(e)}")
            print(f"DEBUG: Error en load_categorias: {e}")
    
    def nueva_categoria(self):
        """Crear nueva categoría"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Nueva Categoría")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="📁 NUEVA CATEGORÍA", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        tk.Label(fields_frame, text="Nombre:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        nombre_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        nombre_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        activo_var = tk.BooleanVar(value=True)
        tk.Checkbutton(fields_frame, text="Categoría activa", variable=activo_var, 
                      font=('Arial', 11), bg='#f8f9fa').grid(row=1, column=1, sticky='w', pady=8)
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        def guardar_categoria():
            nombre = nombre_entry.get().strip()
            activo = activo_var.get()
            
            if not nombre:
                messagebox.showerror("Error", "El nombre es obligatorio")
                nombre_entry.focus()
                return
            
            try:
                query = "INSERT INTO categorias (nombre, activo) VALUES (%s, %s)"
                result = db.execute_one(query, (nombre, activo))
                
                if result:
                    messagebox.showinfo("Éxito", f"Categoría '{nombre}' creada correctamente")
                    dialog.destroy()
                    self.load_categorias()
                else:
                    messagebox.showerror("Error", "No se pudo crear la categoría")
            except Exception as e:
                messagebox.showerror("Error", f"Error guardando categoría: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="GUARDAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=guardar_categoria).pack(side='right', padx=5)
        
        nombre_entry.focus()
    
    def editar_categoria(self):
        """Editar categoría seleccionada"""
        selection = self.categorias_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione una categoría para editar")
            return
        
        categoria_id = self.categorias_tree.item(selection[0])['values'][0]
        
        # Obtener datos actuales
        query = "SELECT * FROM categorias WHERE id = %s"
        categoria = db.execute_one(query, (categoria_id,))
        
        if not categoria:
            messagebox.showerror("Error", "Categoría no encontrada")
            return
        
        # Dialog de edición
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Categoría")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="✏️ EDITAR CATEGORÍA", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        tk.Label(fields_frame, text="Nombre:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        nombre_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        nombre_entry.insert(0, categoria['nombre'])
        nombre_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        activo_var = tk.BooleanVar(value=bool(categoria['activo']))
        tk.Checkbutton(fields_frame, text="Categoría activa", variable=activo_var, 
                      font=('Arial', 11), bg='#f8f9fa').grid(row=1, column=1, sticky='w', pady=8)
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        def actualizar_categoria():
            nombre = nombre_entry.get().strip()
            activo = activo_var.get()
            
            if not nombre:
                messagebox.showerror("Error", "El nombre es obligatorio")
                nombre_entry.focus()
                return
            
            try:
                query = "UPDATE categorias SET nombre = %s, activo = %s WHERE id = %s"
                result = db.execute_query(query, (nombre, activo, categoria_id))
                
                if result:
                    messagebox.showinfo("Éxito", f"Categoría '{nombre}' actualizada correctamente")
                    dialog.destroy()
                    self.load_categorias()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar la categoría")
            except Exception as e:
                messagebox.showerror("Error", f"Error actualizando categoría: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="ACTUALIZAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=actualizar_categoria).pack(side='right', padx=5)
        
        nombre_entry.focus()
    
    def eliminar_categoria(self):
        """Eliminar categoría seleccionada"""
        selection = self.categorias_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione una categoría para eliminar")
            return
        
        categoria_id = self.categorias_tree.item(selection[0])['values'][0]
        categoria_nombre = self.categorias_tree.item(selection[0])['values'][1]
        
        # Verificar si tiene productos
        query = "SELECT COUNT(*) as total FROM productos WHERE categoria_id = %s"
        result = db.execute_one(query, (categoria_id,))
        total_productos = result['total'] if result else 0
        
        if total_productos > 0:
            messagebox.showerror("Error", 
                f"No se puede eliminar la categoría '{categoria_nombre}'\n"
                f"Tiene {total_productos} productos asociados.\n\n"
                f"Primero elimine o cambie la categoría de los productos.")
            return
        
        # Confirmar eliminación
        respuesta = messagebox.askyesno("Confirmar Eliminación", 
            f"¿Está seguro de eliminar la categoría '{categoria_nombre}'?\n\n"
            f"Esta acción no se puede deshacer.")
        
        if respuesta:
            try:
                query = "DELETE FROM categorias WHERE id = %s"
                result = db.execute_query(query, (categoria_id,))
                
                if result:
                    messagebox.showinfo("Éxito", f"Categoría '{categoria_nombre}' eliminada correctamente")
                    self.load_categorias()
                else:
                    messagebox.showerror("Error", "No se pudo eliminar la categoría")
            except Exception as e:
                messagebox.showerror("Error", f"Error eliminando categoría: {str(e)}")
    
    def exportar_categorias(self):
        """Exportar categorías a Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener datos
            query = """
            SELECT c.id, c.nombre, c.activo, COUNT(p.id) as total_productos
            FROM categorias c
            LEFT JOIN productos p ON c.id = p.categoria_id AND p.activo = 1
            GROUP BY c.id, c.nombre, c.activo
            ORDER BY c.nombre
            """
            categorias = db.execute_query(query)
            
            # Crear workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Categorías"
            
            # Headers
            headers = ['ID', 'Nombre', 'Activo', 'Total Productos']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, categoria in enumerate(categorias, 2):
                sheet.cell(row=row, column=1, value=categoria['id'])
                sheet.cell(row=row, column=2, value=categoria['nombre'])
                sheet.cell(row=row, column=3, value='Sí' if categoria['activo'] else 'No')
                sheet.cell(row=row, column=4, value=categoria['total_productos'])
            
            # Ajustar ancho de columnas
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 15
            
            # Guardar archivo
            filename = f"categorias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Categorías exportadas correctamente\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exportando categorías: {str(e)}")
    
    # ================== FUNCIONES DE PRODUCTOS ==================
    
    def load_categoria_filter(self):
        """Cargar opciones de filtro de categorías"""
        try:
            query = "SELECT id, nombre FROM categorias WHERE activo = 1 ORDER BY nombre"
            categorias = db.execute_query(query)
            
            options = ['Todas las categorías']
            for categoria in categorias:
                options.append(f"{categoria['nombre']} (ID: {categoria['id']})")
            
            self.categoria_filter['values'] = options
            self.categoria_filter.set('Todas las categorías')
            
        except Exception as e:
            print(f"Error cargando filtro de categorías: {e}")
    
    def load_productos(self):
        """Cargar productos en el treeview"""
        try:
            # Limpiar tree
            for item in self.productos_tree.get_children():
                self.productos_tree.delete(item)
            
            # Obtener filtro de categoría
            filtro = self.categoria_filter.get()
            categoria_id = None
            
            if filtro != 'Todas las categorías' and 'ID:' in filtro:
                # Extraer ID de la categoría del filtro
                categoria_id = filtro.split('ID: ')[1].rstrip(')')
            
            # Query base (sin descripción para evitar errores)
            query = """
            SELECT p.id, p.nombre, p.categoria_id, p.precio, p.activo, c.nombre as categoria_nombre
            FROM productos p
            JOIN categorias c ON p.categoria_id = c.id
            """
            params = []
            
            if categoria_id:
                query += " WHERE p.categoria_id = %s"
                params.append(categoria_id)
            
            query += " ORDER BY c.nombre, p.nombre"
            
            productos = db.execute_query(query, params)
            
            for producto in productos:
                estado = "✅ Activo" if producto['activo'] else "❌ Inactivo"
                self.productos_tree.insert('', 'end', values=(
                    producto['id'],
                    producto['nombre'],
                    producto['categoria_nombre'],
                    f"${float(producto['precio']):.2f}",
                    estado
                ))
                
        except Exception as e:
            messagebox.showerror("Error", f"Error cargando productos: {str(e)}")
            print(f"DEBUG: Error en load_productos: {e}")
    
    def nuevo_producto(self):
        """Crear nuevo producto"""
        # Primero verificar que hay categorías activas
        query = "SELECT COUNT(*) as total FROM categorias WHERE activo = 1"
        result = db.execute_one(query)
        if not result or result['total'] == 0:
            messagebox.showerror("Error", "No hay categorías activas.\nPrimero cree al menos una categoría.")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Producto")
        dialog.geometry("500x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="🍕 NUEVO PRODUCTO", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        # Nombre
        tk.Label(fields_frame, text="Nombre:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        nombre_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        nombre_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        # Categoría
        tk.Label(fields_frame, text="Categoría:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=1, column=0, sticky='w', pady=8)
        
        categoria_combo = ttk.Combobox(fields_frame, width=28, state='readonly')
        categoria_combo.grid(row=1, column=1, pady=8, padx=10, sticky='ew')
        
        # Cargar categorías
        query = "SELECT id, nombre FROM categorias WHERE activo = 1 ORDER BY nombre"
        categorias = db.execute_query(query)
        categoria_options = []
        categoria_ids = []
        for categoria in categorias:
            categoria_options.append(categoria['nombre'])
            categoria_ids.append(categoria['id'])
        categoria_combo['values'] = categoria_options
        
        # Precio
        tk.Label(fields_frame, text="Precio ($):", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=2, column=0, sticky='w', pady=8)
        precio_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        precio_entry.grid(row=2, column=1, pady=8, padx=10, sticky='ew')
        
        activo_var = tk.BooleanVar(value=True)
        tk.Checkbutton(fields_frame, text="Producto activo", variable=activo_var, 
                      font=('Arial', 11), bg='#f8f9fa').grid(row=3, column=1, sticky='w', pady=8)
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        def guardar_producto():
            nombre = nombre_entry.get().strip()
            precio_str = precio_entry.get().strip()
            activo = activo_var.get()
            categoria_idx = categoria_combo.current()
            
            if not nombre:
                messagebox.showerror("Error", "El nombre es obligatorio")
                nombre_entry.focus()
                return
            
            if categoria_idx < 0:
                messagebox.showerror("Error", "Seleccione una categoría")
                categoria_combo.focus()
                return
            
            try:
                precio = float(precio_str)
                if precio < 0:
                    messagebox.showerror("Error", "El precio no puede ser negativo")
                    precio_entry.focus()
                    return
            except ValueError:
                messagebox.showerror("Error", "Precio inválido")
                precio_entry.focus()
                return
            
            categoria_id = categoria_ids[categoria_idx]
            
            try:
                query = "INSERT INTO productos (nombre, categoria_id, precio, activo) VALUES (%s, %s, %s, %s)"
                result = db.execute_one(query, (nombre, categoria_id, precio, activo))
                
                if result:
                    messagebox.showinfo("Éxito", f"Producto '{nombre}' creado correctamente")
                    dialog.destroy()
                    self.load_productos()
                else:
                    messagebox.showerror("Error", "No se pudo crear el producto")
            except Exception as e:
                messagebox.showerror("Error", f"Error guardando producto: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="GUARDAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=guardar_producto).pack(side='right', padx=5)
        
        nombre_entry.focus()

    def admin_usuarios(self):
        self.show_user_administration()
    
    def show_user_administration(self):
        self.clear_window()
        
        # Header
        header_frame = tk.Frame(self.root, bg='#34495e', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        user_info = f"Admin - Gestión de Usuarios: {auth.current_user['nombre']}"
        tk.Label(header_frame, text=user_info, font=('Arial', 12, 'bold'), 
                bg='#34495e', fg='white').pack(side='left', padx=20, pady=20)
        
        # Botón volver
        tk.Button(header_frame, text="← Volver", font=('Arial', 10), 
                 command=self.show_admin_menu).pack(side='right', padx=20, pady=15)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Título
        tk.Label(main_frame, text="👥 ADMINISTRACIÓN DE USUARIOS", font=('Arial', 18, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Frame para pestañas/secciones
        notebook_frame = tk.Frame(main_frame, bg='#f8f9fa')
        notebook_frame.pack(fill='both', expand=True)
        
        # Botones de sección
        section_frame = tk.Frame(notebook_frame, bg='#f8f9fa')
        section_frame.pack(fill='x', pady=(0,10))
        
        self.current_user_section = tk.StringVar(value='usuarios')
        
        tk.Radiobutton(section_frame, text="👥 USUARIOS", variable=self.current_user_section, 
                      value='usuarios', font=('Arial', 12, 'bold'), bg='#f8f9fa',
                      command=self.refresh_user_content).pack(side='left', padx=20)
        tk.Radiobutton(section_frame, text="🔒 ROLES Y PERMISOS", variable=self.current_user_section, 
                      value='roles', font=('Arial', 12, 'bold'), bg='#f8f9fa',
                      command=self.refresh_user_content).pack(side='left', padx=20)
        tk.Radiobutton(section_frame, text="📊 REPORTES", variable=self.current_user_section, 
                      value='reportes', font=('Arial', 12, 'bold'), bg='#f8f9fa',
                      command=self.refresh_user_content).pack(side='left', padx=20)
        
        # Contenedor para el contenido dinámico
        self.user_content_frame = tk.Frame(notebook_frame, bg='#f8f9fa')
        self.user_content_frame.pack(fill='both', expand=True)
        
        # Cargar contenido inicial
        self.refresh_user_content()
    
    def refresh_user_content(self):
        # Limpiar contenido anterior
        for widget in self.user_content_frame.winfo_children():
            widget.destroy()
        
        section = self.current_user_section.get()
        
        if section == 'usuarios':
            self.show_users_management()
        elif section == 'roles':
            self.show_roles_management()
        elif section == 'reportes':
            self.show_user_reports()
    
    def show_users_management(self):
        # Frame principal para usuarios
        frame = tk.Frame(self.user_content_frame, bg='#f8f9fa')
        frame.pack(fill='both', expand=True)
        
        # Filtros y búsqueda
        filter_frame = tk.Frame(frame, bg='#f8f9fa')
        filter_frame.pack(fill='x', pady=(0,10))
        
        tk.Label(filter_frame, text="🔍 Buscar:", font=('Arial', 11), 
                bg='#f8f9fa').pack(side='left', padx=5)
        
        self.user_search_var = tk.StringVar()
        search_entry = tk.Entry(filter_frame, textvariable=self.user_search_var, font=('Arial', 11), width=25)
        search_entry.pack(side='left', padx=5)
        search_entry.bind('<KeyRelease>', lambda e: self.load_users())
        
        tk.Label(filter_frame, text="Rol:", font=('Arial', 11), 
                bg='#f8f9fa').pack(side='left', padx=(20,5))
        
        self.role_filter = ttk.Combobox(filter_frame, width=15, state='readonly')
        self.role_filter.pack(side='left', padx=5)
        self.role_filter.bind('<<ComboboxSelected>>', lambda e: self.load_users())
        
        # Botones de acción
        action_frame = tk.Frame(frame, bg='#f8f9fa')
        action_frame.pack(fill='x', pady=(0,15))
        
        tk.Button(action_frame, text="➕ NUEVO USUARIO", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=18, height=1,
                 command=self.nuevo_usuario).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="✏️ EDITAR", font=('Arial', 11, 'bold'),
                 bg='#2980b9', fg='white', width=12, height=1,
                 command=self.editar_usuario).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="🔒 CAMBIAR CONTRASEÑA", font=('Arial', 11, 'bold'),
                 bg='#f39c12', fg='white', width=18, height=1,
                 command=self.cambiar_password).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="❌ DESACTIVAR", font=('Arial', 11, 'bold'),
                 bg='#e74c3c', fg='white', width=12, height=1,
                 command=self.toggle_usuario_estado).pack(side='left', padx=5)
        
        tk.Button(action_frame, text="📄 EXPORTAR", font=('Arial', 11, 'bold'),
                 bg='#27ae60', fg='white', width=12, height=1,
                 command=self.exportar_usuarios).pack(side='right', padx=5)
        
        # Lista de usuarios
        list_frame = tk.Frame(frame, bg='#f8f9fa')
        list_frame.pack(fill='both', expand=True)
        
        # Treeview para usuarios
        columns = ('id', 'nombre', 'usuario', 'rol', 'email', 'activo', 'ultimo_acceso')
        self.users_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        self.users_tree.heading('id', text='ID')
        self.users_tree.heading('nombre', text='Nombre Completo')
        self.users_tree.heading('usuario', text='Usuario')
        self.users_tree.heading('rol', text='Rol')
        self.users_tree.heading('email', text='Email')
        self.users_tree.heading('activo', text='Estado')
        self.users_tree.heading('ultimo_acceso', text='Último Acceso')
        
        self.users_tree.column('id', width=50)
        self.users_tree.column('nombre', width=180)
        self.users_tree.column('usuario', width=120)
        self.users_tree.column('rol', width=100)
        self.users_tree.column('email', width=180)
        self.users_tree.column('activo', width=80)
        self.users_tree.column('ultimo_acceso', width=120)
        
        scrollbar_users = ttk.Scrollbar(list_frame, orient='vertical', command=self.users_tree.yview)
        self.users_tree.configure(yscrollcommand=scrollbar_users.set)
        
        self.users_tree.pack(side='left', fill='both', expand=True)
        scrollbar_users.pack(side='right', fill='y')
        
        # Cargar datos
        self.load_role_filter()
        self.load_users()
    
    def show_roles_management(self):
        frame = tk.Frame(self.user_content_frame, bg='#f8f9fa')
        frame.pack(fill='both', expand=True)
        
        # Información de roles del sistema
        info_frame = tk.LabelFrame(frame, text="Roles del Sistema", font=('Arial', 12, 'bold'), 
                                 bg='#f8f9fa', fg='#2c3e50')
        info_frame.pack(fill='x', pady=(0,20), padx=10)
        
        roles_info = [
            ("👑 ADMIN", "Acceso completo al sistema", "#e74c3c"),
            ("👨‍💼 GERENTE", "Gestión de usuarios, reportes, menú", "#f39c12"),
            ("👨‍🍳 CAJERO", "Operación de POS, turnos, órdenes", "#3498db"),
            ("📊 SUPERVISOR", "Supervisión, reportes, cortes", "#9b59b6")
        ]
        
        for i, (rol, descripcion, color) in enumerate(roles_info):
            role_frame = tk.Frame(info_frame, bg='#f8f9fa')
            role_frame.pack(fill='x', padx=10, pady=5)
            
            tk.Label(role_frame, text=rol, font=('Arial', 11, 'bold'), 
                    bg='#f8f9fa', fg=color, width=15, anchor='w').pack(side='left')
            tk.Label(role_frame, text=descripcion, font=('Arial', 11), 
                    bg='#f8f9fa', fg='#2c3e50').pack(side='left', padx=(10,0))
        
        # Matriz de permisos
        permisos_frame = tk.LabelFrame(frame, text="Matriz de Permisos", font=('Arial', 12, 'bold'), 
                                     bg='#f8f9fa', fg='#2c3e50')
        permisos_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Crear matriz de permisos
        self.create_permissions_matrix(permisos_frame)
    
    def create_permissions_matrix(self, parent):
        # Canvas para scroll
        canvas = tk.Canvas(parent, bg='#f8f9fa')
        scrollbar_matrix = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_matrix.set)
        
        # Encabezados
        tk.Label(scrollable_frame, text="FUNCIONALIDAD", font=('Arial', 11, 'bold'), 
                bg='#34495e', fg='white', width=25).grid(row=0, column=0, sticky='ew', padx=1, pady=1)
        
        roles = ["ADMIN", "GERENTE", "CAJERO", "SUPERVISOR"]
        for i, rol in enumerate(roles):
            tk.Label(scrollable_frame, text=rol, font=('Arial', 10, 'bold'), 
                    bg='#34495e', fg='white', width=12).grid(row=0, column=i+1, sticky='ew', padx=1, pady=1)
        
        # Definir permisos por funcionalidad
        permisos = [
            ("POS - Crear Órdenes", ["❌", "✅", "✅", "✅"]),
            ("POS - Procesar Pagos", ["❌", "✅", "✅", "✅"]),
            ("POS - Cancelar Órdenes", ["✅", "✅", "❌", "✅"]),
            ("Turnos - Abrir/Cerrar", ["✅", "✅", "✅", "❌"]),
            ("Arqueos de Caja", ["✅", "✅", "✅", "✅"]),
            ("Cortes de Caja", ["✅", "✅", "✅", "✅"]),
            ("Historial de Órdenes", ["✅", "✅", "✅", "✅"]),
            ("Administrar Menú", ["✅", "✅", "❌", "❌"]),
            ("Administrar Usuarios", ["✅", "✅", "❌", "❌"]),
            ("Reportes de Ventas", ["✅", "✅", "❌", "✅"]),
            ("Reportes de Usuarios", ["✅", "✅", "❌", "❌"]),
            ("Configuración Sistema", ["✅", "❌", "❌", "❌"])
        ]
        
        for row, (funcionalidad, perms) in enumerate(permisos, 1):
            # Funcionalidad
            tk.Label(scrollable_frame, text=funcionalidad, font=('Arial', 10), 
                    bg='#ecf0f1', fg='#2c3e50', width=25, anchor='w').grid(row=row, column=0, sticky='ew', padx=1, pady=1)
            
            # Permisos por rol
            for col, permiso in enumerate(perms):
                color = '#27ae60' if permiso == '✅' else '#e74c3c'
                tk.Label(scrollable_frame, text=permiso, font=('Arial', 12, 'bold'), 
                        bg='#ffffff', fg=color, width=12).grid(row=row, column=col+1, sticky='ew', padx=1, pady=1)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_matrix.pack(side="right", fill="y")
    
    def show_user_reports(self):
        frame = tk.Frame(self.user_content_frame, bg='#f8f9fa')
        frame.pack(fill='both', expand=True)
        
        # Título de reportes
        tk.Label(frame, text="📊 REPORTES DE USUARIOS", font=('Arial', 16, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Grid de botones de reportes
        reports_grid = tk.Frame(frame, bg='#f8f9fa')
        reports_grid.pack(expand=True)
        
        # Fila 1
        row1 = tk.Frame(reports_grid, bg='#f8f9fa')
        row1.pack(pady=10)
        
        tk.Button(row1, text="👥 LISTADO COMPLETO", font=('Arial', 12, 'bold'),
                 bg='#3498db', fg='white', width=25, height=3,
                 command=self.reporte_usuarios_completo).pack(side='left', padx=10)
        
        tk.Button(row1, text="🔐 ACTIVIDAD DE ACCESOS", font=('Arial', 12, 'bold'),
                 bg='#2980b9', fg='white', width=25, height=3,
                 command=self.reporte_actividad_accesos).pack(side='left', padx=10)
        
        # Fila 2
        row2 = tk.Frame(reports_grid, bg='#f8f9fa')
        row2.pack(pady=10)
        
        tk.Button(row2, text="📈 USUARIOS POR ROL", font=('Arial', 12, 'bold'),
                 bg='#5dade2', fg='white', width=25, height=3,
                 command=self.reporte_usuarios_por_rol).pack(side='left', padx=10)
        
        tk.Button(row2, text="⚠️ USUARIOS INACTIVOS", font=('Arial', 12, 'bold'),
                 bg='#e74c3c', fg='white', width=25, height=3,
                 command=self.reporte_usuarios_inactivos).pack(side='left', padx=10)

    # ================== FUNCIONES DE GESTIÓN DE USUARIOS ==================
    
    def load_role_filter(self):
        """Cargar opciones de filtro de roles"""
        try:
            roles = ['Todos los roles', 'admin', 'gerente', 'cajero', 'supervisor']
            self.role_filter['values'] = roles
            self.role_filter.set('Todos los roles')
        except Exception as e:
            print(f"Error cargando filtro de roles: {e}")
    
    def load_users(self):
        """Cargar usuarios en el treeview"""
        try:
            # Limpiar tree
            for item in self.users_tree.get_children():
                self.users_tree.delete(item)
            
            # Obtener filtros
            search_text = self.user_search_var.get().lower()
            role_filter = self.role_filter.get()
            
            # Query base
            query = """
            SELECT u.id, u.nombre, u.usuario, u.rol, u.email, u.activo, 
                   u.ultimo_acceso, u.fecha_creacion
            FROM usuarios u
            WHERE 1=1
            """
            params = []
            
            # Aplicar filtro de búsqueda
            if search_text:
                query += " AND (LOWER(u.nombre) LIKE %s OR LOWER(u.usuario) LIKE %s OR LOWER(u.email) LIKE %s)"
                search_param = f"%{search_text}%"
                params.extend([search_param, search_param, search_param])
            
            # Aplicar filtro de rol
            if role_filter != 'Todos los roles':
                query += " AND u.rol = %s"
                params.append(role_filter)
            
            query += " ORDER BY u.activo DESC, u.nombre"
            
            usuarios = db.execute_query(query, params)
            
            for usuario in usuarios:
                estado = "✅ Activo" if usuario['activo'] else "❌ Inactivo"
                ultimo_acceso = ""
                if usuario['ultimo_acceso']:
                    ultimo_acceso = usuario['ultimo_acceso'].strftime('%d/%m/%Y %H:%M')
                
                # Determinar color de rol
                rol_display = usuario['rol'].upper()
                
                self.users_tree.insert('', 'end', values=(
                    usuario['id'],
                    usuario['nombre'],
                    usuario['usuario'],
                    rol_display,
                    usuario['email'] or '',
                    estado,
                    ultimo_acceso
                ))
                
        except Exception as e:
            messagebox.showerror("Error", f"Error cargando usuarios: {str(e)}")
            print(f"DEBUG: Error en load_users: {e}")
    
    def nuevo_usuario(self):
        """Crear nuevo usuario"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Usuario")
        dialog.geometry("500x600")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="👥 NUEVO USUARIO", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        # Nombre completo
        tk.Label(fields_frame, text="Nombre Completo:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        nombre_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        nombre_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        # Usuario (login)
        tk.Label(fields_frame, text="Usuario (login):", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=1, column=0, sticky='w', pady=8)
        usuario_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        usuario_entry.grid(row=1, column=1, pady=8, padx=10, sticky='ew')
        
        # Email
        tk.Label(fields_frame, text="Email:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=2, column=0, sticky='w', pady=8)
        email_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        email_entry.grid(row=2, column=1, pady=8, padx=10, sticky='ew')
        
        # Rol
        tk.Label(fields_frame, text="Rol:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=3, column=0, sticky='w', pady=8)
        rol_combo = ttk.Combobox(fields_frame, width=28, state='readonly')
        rol_combo['values'] = ('cajero', 'supervisor', 'gerente', 'admin')
        rol_combo.set('cajero')
        rol_combo.grid(row=3, column=1, pady=8, padx=10, sticky='ew')
        
        # Contraseña
        tk.Label(fields_frame, text="Contraseña:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=4, column=0, sticky='w', pady=8)
        password_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1, show='*')
        password_entry.grid(row=4, column=1, pady=8, padx=10, sticky='ew')
        
        # Confirmar contraseña
        tk.Label(fields_frame, text="Confirmar Contraseña:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=5, column=0, sticky='w', pady=8)
        confirm_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1, show='*')
        confirm_entry.grid(row=5, column=1, pady=8, padx=10, sticky='ew')
        
        # Estado activo
        activo_var = tk.BooleanVar(value=True)
        tk.Checkbutton(fields_frame, text="Usuario activo", variable=activo_var, 
                      font=('Arial', 11), bg='#f8f9fa').grid(row=6, column=1, sticky='w', pady=8)
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        # Información de seguridad
        security_frame = tk.LabelFrame(main_frame, text="Política de Contraseñas", font=('Arial', 11), 
                                     bg='#f8f9fa', fg='#2c3e50')
        security_frame.pack(fill='x', pady=(0,20))
        
        security_info = [
            "• Mínimo 6 caracteres",
            "• Se recomienda usar números y letras",
            "• Evitar contraseñas obvias como '123456'",
            "• El usuario deberá cambiarla en el primer acceso"
        ]
        
        for info in security_info:
            tk.Label(security_frame, text=info, font=('Arial', 9), 
                    bg='#f8f9fa', fg='#7f8c8d').pack(anchor='w', padx=10, pady=2)
        
        def guardar_usuario():
            nombre = nombre_entry.get().strip()
            usuario = usuario_entry.get().strip()
            email = email_entry.get().strip()
            rol = rol_combo.get()
            password = password_entry.get()
            confirm = confirm_entry.get()
            activo = activo_var.get()
            
            # Validaciones
            if not nombre:
                messagebox.showerror("Error", "El nombre completo es obligatorio")
                nombre_entry.focus()
                return
            
            if not usuario:
                messagebox.showerror("Error", "El usuario es obligatorio")
                usuario_entry.focus()
                return
            
            if len(usuario) < 3:
                messagebox.showerror("Error", "El usuario debe tener al menos 3 caracteres")
                usuario_entry.focus()
                return
            
            if not password:
                messagebox.showerror("Error", "La contraseña es obligatoria")
                password_entry.focus()
                return
            
            if len(password) < 6:
                messagebox.showerror("Error", "La contraseña debe tener al menos 6 caracteres")
                password_entry.focus()
                return
            
            if password != confirm:
                messagebox.showerror("Error", "Las contraseñas no coinciden")
                confirm_entry.focus()
                return
            
            # Validar email si se proporciona
            if email and '@' not in email:
                messagebox.showerror("Error", "Email inválido")
                email_entry.focus()
                return
            
            try:
                # Verificar si el usuario ya existe
                check_query = "SELECT COUNT(*) as count FROM usuarios WHERE usuario = %s"
                result = db.execute_one(check_query, (usuario,))
                if result and result['count'] > 0:
                    messagebox.showerror("Error", f"El usuario '{usuario}' ya existe")
                    usuario_entry.focus()
                    return
                
                # Hashear contraseña (simple hash - en producción usar bcrypt)
                import hashlib
                password_hash = hashlib.sha256(password.encode()).hexdigest()
                
                # Insertar usuario
                query = """
                INSERT INTO usuarios (nombre, usuario, password, email, rol, activo) 
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                result = db.execute_one(query, (nombre, usuario, password_hash, email or None, rol, activo))
                
                if result:
                    messagebox.showinfo("Éxito", f"Usuario '{nombre}' creado correctamente")
                    dialog.destroy()
                    self.load_users()
                else:
                    messagebox.showerror("Error", "No se pudo crear el usuario")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Error guardando usuario: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="GUARDAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=guardar_usuario).pack(side='right', padx=5)
        
        nombre_entry.focus()
    
    def editar_usuario(self):
        """Editar usuario seleccionado"""
        selection = self.users_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un usuario para editar")
            return
        
        user_id = self.users_tree.item(selection[0])['values'][0]
        
        # No permitir editar el usuario actual
        if user_id == auth.current_user['id']:
            messagebox.showwarning("Advertencia", "No puede editar su propio usuario desde aquí")
            return
        
        # Obtener datos del usuario
        query = "SELECT * FROM usuarios WHERE id = %s"
        usuario = db.execute_one(query, (user_id,))
        
        if not usuario:
            messagebox.showerror("Error", "Usuario no encontrado")
            return
        
        # Dialog de edición
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Usuario")
        dialog.geometry("500x500")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="✏️ EDITAR USUARIO", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Información del usuario
        info_frame = tk.LabelFrame(main_frame, text="Información Actual", font=('Arial', 11), 
                                 bg='#f8f9fa', fg='#2c3e50')
        info_frame.pack(fill='x', pady=(0,15))
        
        tk.Label(info_frame, text=f"Usuario: {usuario['usuario']}", font=('Arial', 10), 
                bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
        tk.Label(info_frame, text=f"Creado: {usuario['fecha_creacion'].strftime('%d/%m/%Y')}", 
                font=('Arial', 10), bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
        
        # Campos editables
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        # Nombre completo
        tk.Label(fields_frame, text="Nombre Completo:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        nombre_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        nombre_entry.insert(0, usuario['nombre'])
        nombre_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        # Email
        tk.Label(fields_frame, text="Email:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=1, column=0, sticky='w', pady=8)
        email_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        email_entry.insert(0, usuario['email'] or '')
        email_entry.grid(row=1, column=1, pady=8, padx=10, sticky='ew')
        
        # Rol
        tk.Label(fields_frame, text="Rol:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=2, column=0, sticky='w', pady=8)
        rol_combo = ttk.Combobox(fields_frame, width=28, state='readonly')
        rol_combo['values'] = ('cajero', 'supervisor', 'gerente', 'admin')
        rol_combo.set(usuario['rol'])
        rol_combo.grid(row=2, column=1, pady=8, padx=10, sticky='ew')
        
        # Estado activo
        activo_var = tk.BooleanVar(value=bool(usuario['activo']))
        tk.Checkbutton(fields_frame, text="Usuario activo", variable=activo_var, 
                      font=('Arial', 11), bg='#f8f9fa').grid(row=3, column=1, sticky='w', pady=8)
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        def actualizar_usuario():
            nombre = nombre_entry.get().strip()
            email = email_entry.get().strip()
            rol = rol_combo.get()
            activo = activo_var.get()
            
            # Validaciones
            if not nombre:
                messagebox.showerror("Error", "El nombre completo es obligatorio")
                nombre_entry.focus()
                return
            
            if email and '@' not in email:
                messagebox.showerror("Error", "Email inválido")
                email_entry.focus()
                return
            
            try:
                query = "UPDATE usuarios SET nombre = %s, email = %s, rol = %s, activo = %s WHERE id = %s"
                result = db.execute_query(query, (nombre, email or None, rol, activo, user_id))
                
                if result:
                    messagebox.showinfo("Éxito", f"Usuario '{nombre}' actualizado correctamente")
                    dialog.destroy()
                    self.load_users()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar el usuario")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Error actualizando usuario: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="ACTUALIZAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=actualizar_usuario).pack(side='right', padx=5)
        
        nombre_entry.focus()
    
    def cambiar_password(self):
        """Cambiar contraseña de usuario"""
        selection = self.users_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un usuario para cambiar contraseña")
            return
        
        user_id = self.users_tree.item(selection[0])['values'][0]
        user_name = self.users_tree.item(selection[0])['values'][1]
        
        # Dialog para cambio de contraseña
        dialog = tk.Toplevel(self.root)
        dialog.title("Cambiar Contraseña")
        dialog.geometry("400x350")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="🔒 CAMBIAR CONTRASEÑA", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        tk.Label(main_frame, text=f"Usuario: {user_name}", font=('Arial', 12), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        # Nueva contraseña
        tk.Label(fields_frame, text="Nueva Contraseña:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        password_entry = tk.Entry(fields_frame, font=('Arial', 11), width=25, relief='solid', bd=1, show='*')
        password_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        # Confirmar contraseña
        tk.Label(fields_frame, text="Confirmar Contraseña:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=1, column=0, sticky='w', pady=8)
        confirm_entry = tk.Entry(fields_frame, font=('Arial', 11), width=25, relief='solid', bd=1, show='*')
        confirm_entry.grid(row=1, column=1, pady=8, padx=10, sticky='ew')
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        # Política de contraseñas
        policy_frame = tk.LabelFrame(main_frame, text="Política de Contraseñas", font=('Arial', 10), 
                                   bg='#f8f9fa', fg='#2c3e50')
        policy_frame.pack(fill='x', pady=(0,15))
        
        tk.Label(policy_frame, text="• Mínimo 6 caracteres", font=('Arial', 9), 
                bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
        tk.Label(policy_frame, text="• Se recomienda usar números y letras", font=('Arial', 9), 
                bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
        
        def cambiar():
            password = password_entry.get()
            confirm = confirm_entry.get()
            
            if not password:
                messagebox.showerror("Error", "La contraseña es obligatoria")
                password_entry.focus()
                return
            
            if len(password) < 6:
                messagebox.showerror("Error", "La contraseña debe tener al menos 6 caracteres")
                password_entry.focus()
                return
            
            if password != confirm:
                messagebox.showerror("Error", "Las contraseñas no coinciden")
                confirm_entry.focus()
                return
            
            try:
                import hashlib
                password_hash = hashlib.sha256(password.encode()).hexdigest()
                
                query = "UPDATE usuarios SET password = %s WHERE id = %s"
                result = db.execute_query(query, (password_hash, user_id))
                
                if result:
                    messagebox.showinfo("Éxito", f"Contraseña cambiada para {user_name}")
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "No se pudo cambiar la contraseña")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Error cambiando contraseña: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="CAMBIAR", font=('Arial', 11, 'bold'),
                 bg='#f39c12', fg='white', width=12, height=1,
                 command=cambiar).pack(side='right', padx=5)
        
        password_entry.focus()
    
    def toggle_usuario_estado(self):
        """Activar/Desactivar usuario"""
        selection = self.users_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un usuario para cambiar estado")
            return
        
        user_id = self.users_tree.item(selection[0])['values'][0]
        user_name = self.users_tree.item(selection[0])['values'][1]
        current_status = self.users_tree.item(selection[0])['values'][5]
        
        # No permitir desactivar el usuario actual
        if user_id == auth.current_user['id']:
            messagebox.showwarning("Advertencia", "No puede desactivar su propio usuario")
            return
        
        # Determinar nueva acción
        is_active = "Activo" in current_status
        action = "desactivar" if is_active else "activar"
        new_status = not is_active
        
        # Confirmar acción
        respuesta = messagebox.askyesno("Confirmar Acción", 
            f"¿Está seguro de {action} el usuario '{user_name}'?\n\n"
            f"Usuario {'no podrá' if not new_status else 'podrá'} acceder al sistema.")
        
        if respuesta:
            try:
                query = "UPDATE usuarios SET activo = %s WHERE id = %s"
                result = db.execute_query(query, (new_status, user_id))
                
                if result:
                    status_text = "activado" if new_status else "desactivado"
                    messagebox.showinfo("Éxito", f"Usuario '{user_name}' {status_text} correctamente")
                    self.load_users()
                else:
                    messagebox.showerror("Error", f"No se pudo {action} el usuario")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Error cambiando estado: {str(e)}")
    
    def exportar_usuarios(self):
        """Exportar usuarios a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener datos
            query = """
            SELECT u.id, u.nombre, u.usuario, u.rol, u.email, u.activo, 
                   u.ultimo_acceso, u.fecha_creacion
            FROM usuarios u
            ORDER BY u.rol, u.nombre
            """
            usuarios = db.execute_query(query)
            
            # Crear workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Usuarios del Sistema"
            
            # Título
            sheet.merge_cells('A1:H1')
            title_cell = sheet['A1']
            title_cell.value = f"LISTADO DE USUARIOS - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill('solid', start_color='2980B9')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['ID', 'Nombre Completo', 'Usuario', 'Rol', 'Email', 'Estado', 'Último Acceso', 'Fecha Creación']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, usuario in enumerate(usuarios, 4):
                sheet.cell(row=row, column=1, value=usuario['id'])
                sheet.cell(row=row, column=2, value=usuario['nombre'])
                sheet.cell(row=row, column=3, value=usuario['usuario'])
                sheet.cell(row=row, column=4, value=usuario['rol'].upper())
                sheet.cell(row=row, column=5, value=usuario['email'] or '')
                sheet.cell(row=row, column=6, value='Activo' if usuario['activo'] else 'Inactivo')
                
                # Formatear fechas
                if usuario['ultimo_acceso']:
                    sheet.cell(row=row, column=7, value=usuario['ultimo_acceso'].strftime('%d/%m/%Y %H:%M'))
                sheet.cell(row=row, column=8, value=usuario['fecha_creacion'].strftime('%d/%m/%Y'))
            
            # Ajustar columnas
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 25
            sheet.column_dimensions['F'].width = 12
            sheet.column_dimensions['G'].width = 18
            sheet.column_dimensions['H'].width = 15
            
            # Guardar archivo
            filename = f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Usuarios exportados correctamente\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exportando usuarios: {str(e)}")
    
    # ================== FUNCIONES DE REPORTES DE USUARIOS ==================
    
    def reporte_usuarios_completo(self):
        """Generar reporte completo de usuarios"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener datos estadísticos
            stats_query = """
            SELECT 
                COUNT(*) as total_usuarios,
                SUM(CASE WHEN activo = 1 THEN 1 ELSE 0 END) as usuarios_activos,
                SUM(CASE WHEN rol = 'admin' THEN 1 ELSE 0 END) as admins,
                SUM(CASE WHEN rol = 'gerente' THEN 1 ELSE 0 END) as gerentes,
                SUM(CASE WHEN rol = 'cajero' THEN 1 ELSE 0 END) as cajeros,
                SUM(CASE WHEN rol = 'supervisor' THEN 1 ELSE 0 END) as supervisores,
                SUM(CASE WHEN ultimo_acceso >= DATE_SUB(NOW(), INTERVAL 30 DAY) THEN 1 ELSE 0 END) as activos_30dias
            FROM usuarios
            """
            stats = db.execute_one(stats_query)
            
            # Obtener usuarios detallados
            users_query = """
            SELECT u.id, u.nombre, u.usuario, u.rol, u.email, u.activo, 
                   u.ultimo_acceso, u.fecha_creacion,
                   DATEDIFF(NOW(), u.ultimo_acceso) as dias_sin_acceso
            FROM usuarios u
            ORDER BY u.rol, u.activo DESC, u.nombre
            """
            usuarios = db.execute_query(users_query)
            
            # Crear workbook
            wb = Workbook()
            
            # Hoja de resumen
            sheet_resumen = wb.active
            sheet_resumen.title = "Resumen Ejecutivo"
            
            # Título resumen
            sheet_resumen.merge_cells('A1:B1')
            title_cell = sheet_resumen['A1']
            title_cell.value = f"REPORTE COMPLETO DE USUARIOS - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill('solid', start_color='2980B9')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Estadísticas generales
            estadisticas = [
                ("Total de usuarios:", stats['total_usuarios']),
                ("Usuarios activos:", stats['usuarios_activos']),
                ("Usuarios inactivos:", stats['total_usuarios'] - stats['usuarios_activos']),
                ("", ""),
                ("Por rol:", ""),
                ("  Administradores:", stats['admins']),
                ("  Gerentes:", stats['gerentes']),
                ("  Cajeros:", stats['cajeros']),
                ("  Supervisores:", stats['supervisores']),
                ("", ""),
                ("Usuarios activos (30 días):", stats['activos_30dias'])
            ]
            
            for row, (label, value) in enumerate(estadisticas, 3):
                if label:
                    sheet_resumen.cell(row=row, column=1, value=label).font = Font(bold=True)
                    sheet_resumen.cell(row=row, column=2, value=value)
            
            # Ajustar columnas del resumen
            sheet_resumen.column_dimensions['A'].width = 25
            sheet_resumen.column_dimensions['B'].width = 15
            
            # Hoja de usuarios detallados
            sheet_usuarios = wb.create_sheet("Usuarios Detallados")
            
            # Headers de usuarios
            headers = ['ID', 'Nombre', 'Usuario', 'Rol', 'Email', 'Estado', 'Último Acceso', 'Días sin Acceso', 'Fecha Creación']
            for col, header in enumerate(headers, 1):
                cell = sheet_usuarios.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de usuarios
            for row, usuario in enumerate(usuarios, 2):
                sheet_usuarios.cell(row=row, column=1, value=usuario['id'])
                sheet_usuarios.cell(row=row, column=2, value=usuario['nombre'])
                sheet_usuarios.cell(row=row, column=3, value=usuario['usuario'])
                sheet_usuarios.cell(row=row, column=4, value=usuario['rol'].upper())
                sheet_usuarios.cell(row=row, column=5, value=usuario['email'] or '')
                sheet_usuarios.cell(row=row, column=6, value='Activo' if usuario['activo'] else 'Inactivo')
                
                if usuario['ultimo_acceso']:
                    sheet_usuarios.cell(row=row, column=7, value=usuario['ultimo_acceso'].strftime('%d/%m/%Y %H:%M'))
                    sheet_usuarios.cell(row=row, column=8, value=usuario['dias_sin_acceso'] or 0)
                else:
                    sheet_usuarios.cell(row=row, column=7, value='Nunca')
                    sheet_usuarios.cell(row=row, column=8, value='N/A')
                
                sheet_usuarios.cell(row=row, column=9, value=usuario['fecha_creacion'].strftime('%d/%m/%Y'))
            
            # Ajustar columnas de usuarios
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                sheet_usuarios.column_dimensions[col].width = 18
            
            # Guardar
            filename = f"reporte_usuarios_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Reporte completo generado\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando reporte: {str(e)}")
    
    def reporte_actividad_accesos(self):
        """Reporte de actividad de accesos"""
        messagebox.showinfo("Función de Actividad", 
            "Reporte de actividad de accesos:\n\n"
            "• Últimos accesos por usuario\n"
            "• Frecuencia de uso del sistema\n"
            "• Usuarios sin actividad reciente\n"
            "• Patrones de horarios de trabajo\n\n"
            "Esta función estará disponible próximamente.")
    
    def reporte_usuarios_por_rol(self):
        """Reporte de usuarios agrupados por rol"""
        messagebox.showinfo("Función de Usuarios por Rol", 
            "Reporte de usuarios por rol:\n\n"
            "• Distribución por tipo de rol\n"
            "• Permisos asignados\n"
            "• Usuarios activos/inactivos por rol\n"
            "• Análisis de estructura organizacional\n\n"
            "Esta función estará disponible próximamente.")
    
    def reporte_usuarios_inactivos(self):
        """Reporte de usuarios inactivos"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener usuarios inactivos
            query = """
            SELECT u.id, u.nombre, u.usuario, u.rol, u.email, 
                   u.ultimo_acceso, u.fecha_creacion,
                   DATEDIFF(NOW(), u.ultimo_acceso) as dias_sin_acceso
            FROM usuarios u
            WHERE u.activo = 0 OR u.ultimo_acceso IS NULL OR u.ultimo_acceso < DATE_SUB(NOW(), INTERVAL 60 DAY)
            ORDER BY u.ultimo_acceso ASC
            """
            usuarios_inactivos = db.execute_query(query)
            
            if not usuarios_inactivos:
                messagebox.showinfo("Información", "No hay usuarios inactivos para reportar")
                return
            
            # Crear Excel
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Usuarios Inactivos"
            
            # Título
            sheet.merge_cells('A1:H1')
            title_cell = sheet['A1']
            title_cell.value = f"USUARIOS INACTIVOS - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill('solid', start_color='E74C3C')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['ID', 'Nombre', 'Usuario', 'Rol', 'Email', 'Último Acceso', 'Días sin Acceso', 'Fecha Creación']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='C0392B')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, usuario in enumerate(usuarios_inactivos, 4):
                sheet.cell(row=row, column=1, value=usuario['id'])
                sheet.cell(row=row, column=2, value=usuario['nombre'])
                sheet.cell(row=row, column=3, value=usuario['usuario'])
                sheet.cell(row=row, column=4, value=usuario['rol'].upper())
                sheet.cell(row=row, column=5, value=usuario['email'] or '')
                
                if usuario['ultimo_acceso']:
                    sheet.cell(row=row, column=6, value=usuario['ultimo_acceso'].strftime('%d/%m/%Y %H:%M'))
                    sheet.cell(row=row, column=7, value=usuario['dias_sin_acceso'])
                else:
                    sheet.cell(row=row, column=6, value='Nunca')
                    sheet.cell(row=row, column=7, value='N/A')
                
                sheet.cell(row=row, column=8, value=usuario['fecha_creacion'].strftime('%d/%m/%Y'))
            
            # Ajustar columnas
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                sheet.column_dimensions[col].width = 18
            
            # Guardar
            filename = f"usuarios_inactivos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Reporte de usuarios inactivos generado\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando reporte: {str(e)}")

    def admin_ventas(self):
        messagebox.showinfo("Información", "Funcionalidad en desarrollo")

    def admin_clientes(self):
        messagebox.showinfo("Información", "Funcionalidad en desarrollo")
        
    # ================== FUNCIONES ADICIONALES DE PRODUCTOS ==================
    
    def editar_producto(self):
        """Editar producto seleccionado"""
        selection = self.productos_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un producto para editar")
            return
        
        producto_id = self.productos_tree.item(selection[0])['values'][0]
        
        # Obtener datos actuales
        query = "SELECT * FROM productos WHERE id = %s"
        producto = db.execute_one(query, (producto_id,))
        
        if not producto:
            messagebox.showerror("Error", "Producto no encontrado")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Producto")
        dialog.geometry("500x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        
        # Frame principal
        main_frame = tk.Frame(dialog, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(main_frame, text="✏️ EDITAR PRODUCTO", font=('Arial', 14, 'bold'), 
                bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
        
        # Campos
        fields_frame = tk.Frame(main_frame, bg='#f8f9fa')
        fields_frame.pack(fill='x', pady=(0,20))
        
        # Nombre
        tk.Label(fields_frame, text="Nombre:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=0, column=0, sticky='w', pady=8)
        nombre_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        nombre_entry.insert(0, producto['nombre'])
        nombre_entry.grid(row=0, column=1, pady=8, padx=10, sticky='ew')
        
        # Categoría
        tk.Label(fields_frame, text="Categoría:", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=1, column=0, sticky='w', pady=8)
        
        categoria_combo = ttk.Combobox(fields_frame, width=28, state='readonly')
        categoria_combo.grid(row=1, column=1, pady=8, padx=10, sticky='ew')
        
        # Cargar categorías
        query = "SELECT id, nombre FROM categorias WHERE activo = 1 ORDER BY nombre"
        categorias = db.execute_query(query)
        categoria_options = []
        categoria_ids = []
        current_idx = 0
        for i, categoria in enumerate(categorias):
            categoria_options.append(categoria['nombre'])
            categoria_ids.append(categoria['id'])
            if categoria['id'] == producto['categoria_id']:
                current_idx = i
        
        categoria_combo['values'] = categoria_options
        categoria_combo.current(current_idx)
        
        # Precio
        tk.Label(fields_frame, text="Precio ($):", font=('Arial', 11), 
                bg='#f8f9fa').grid(row=2, column=0, sticky='w', pady=8)
        precio_entry = tk.Entry(fields_frame, font=('Arial', 11), width=30, relief='solid', bd=1)
        precio_entry.insert(0, str(float(producto['precio'])))
        precio_entry.grid(row=2, column=1, pady=8, padx=10, sticky='ew')
        
        activo_var = tk.BooleanVar(value=bool(producto['activo']))
        tk.Checkbutton(fields_frame, text="Producto activo", variable=activo_var, 
                      font=('Arial', 11), bg='#f8f9fa').grid(row=3, column=1, sticky='w', pady=8)
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
        def actualizar_producto():
            nombre = nombre_entry.get().strip()
            precio_str = precio_entry.get().strip()
            activo = activo_var.get()
            categoria_idx = categoria_combo.current()
            
            if not nombre:
                messagebox.showerror("Error", "El nombre es obligatorio")
                nombre_entry.focus()
                return
            
            if categoria_idx < 0:
                messagebox.showerror("Error", "Seleccione una categoría")
                categoria_combo.focus()
                return
            
            try:
                precio = float(precio_str)
                if precio < 0:
                    messagebox.showerror("Error", "El precio no puede ser negativo")
                    precio_entry.focus()
                    return
            except ValueError:
                messagebox.showerror("Error", "Precio inválido")
                precio_entry.focus()
                return
            
            categoria_id = categoria_ids[categoria_idx]
            
            try:
                query = "UPDATE productos SET nombre = %s, categoria_id = %s, precio = %s, activo = %s WHERE id = %s"
                result = db.execute_query(query, (nombre, categoria_id, precio, activo, producto_id))
                
                if result:
                    messagebox.showinfo("Éxito", f"Producto '{nombre}' actualizado correctamente")
                    dialog.destroy()
                    self.load_productos()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar el producto")
            except Exception as e:
                messagebox.showerror("Error", f"Error actualizando producto: {str(e)}")
        
        # Botones
        button_frame = tk.Frame(main_frame, bg='#f8f9fa')
        button_frame.pack(fill='x')
        
        tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                 bg='#95a5a6', fg='white', width=12, height=1,
                 command=dialog.destroy).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="ACTUALIZAR", font=('Arial', 11, 'bold'),
                 bg='#3498db', fg='white', width=12, height=1,
                 command=actualizar_producto).pack(side='right', padx=5)
        
        nombre_entry.focus()
    
    def eliminar_producto(self):
        """Eliminar producto seleccionado"""
        selection = self.productos_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un producto para eliminar")
            return
        
        producto_id = self.productos_tree.item(selection[0])['values'][0]
        producto_nombre = self.productos_tree.item(selection[0])['values'][1]
        
        # Verificar si se ha usado en órdenes
        query = "SELECT COUNT(*) as total FROM orden_detalles WHERE producto_id = %s"
        result = db.execute_one(query, (producto_id,))
        total_ordenes = result['total'] if result else 0
        
        mensaje = f"¿Está seguro de eliminar el producto '{producto_nombre}'?\n\n"
        
        if total_ordenes > 0:
            mensaje += f"⚠️ ADVERTENCIA: Este producto aparece en {total_ordenes} órdenes.\n"
            mensaje += "Al eliminarlo no aparecerá en reportes futuros, pero las órdenes\n"
            mensaje += "existentes mantendrán la información.\n\n"
        
        mensaje += "Esta acción no se puede deshacer."
        
        respuesta = messagebox.askyesno("Confirmar Eliminación", mensaje)
        
        if respuesta:
            try:
                query = "DELETE FROM productos WHERE id = %s"
                result = db.execute_query(query, (producto_id,))
                
                if result:
                    messagebox.showinfo("Éxito", f"Producto '{producto_nombre}' eliminado correctamente")
                    self.load_productos()
                else:
                    messagebox.showerror("Error", "No se pudo eliminar el producto")
            except Exception as e:
                messagebox.showerror("Error", f"Error eliminando producto: {str(e)}")
    
    def exportar_productos(self):
        """Exportar productos a Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            
            # Obtener datos
            query = """
            SELECT p.id, p.nombre, c.nombre as categoria, p.precio, p.activo
            FROM productos p
            JOIN categorias c ON p.categoria_id = c.id
            ORDER BY c.nombre, p.nombre
            """
            productos = db.execute_query(query)
            
            # Crear workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Productos"
            
            # Headers
            headers = ['ID', 'Nombre', 'Categoría', 'Precio', 'Activo']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, producto in enumerate(productos, 2):
                sheet.cell(row=row, column=1, value=producto['id'])
                sheet.cell(row=row, column=2, value=producto['nombre'])
                sheet.cell(row=row, column=3, value=producto['categoria'])
                sheet.cell(row=row, column=4, value=float(producto['precio']))
                sheet.cell(row=row, column=5, value='Sí' if producto['activo'] else 'No')
            
            # Formatear precios
            for row in range(2, len(productos) + 2):
                cell = sheet.cell(row=row, column=4)
                cell.number_format = '$#,##0.00'
            
            # Ajustar ancho de columnas
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 10
            
            # Guardar archivo
            filename = f"productos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Productos exportados correctamente\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exportando productos: {str(e)}")
    
    def importar_productos(self):
        """Importar productos desde Excel"""
        messagebox.showinfo("Función de Importar", 
            "Funcionalidad de importación desde Excel:\n\n"
            "1. El archivo Excel debe contener las columnas:\n"
            "   - nombre (obligatorio)\n"
            "   - categoria (obligatorio - debe existir)\n"
            "   - precio (obligatorio - número positivo)\n\n"
            "2. Use 'Exportar Productos' para ver el formato correcto\n\n"
            "Esta función estará disponible próximamente.")

    def reporte_productos_vendidos(self):
        """Generar reporte de productos más vendidos"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime, timedelta
            
            # Dialog para seleccionar período
            dialog = tk.Toplevel(self.root)
            dialog.title("Reporte Productos Más Vendidos")
            dialog.geometry("400x300")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.configure(bg='#f8f9fa')
            
            main_frame = tk.Frame(dialog, bg='#f8f9fa')
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            tk.Label(main_frame, text="📈 PRODUCTOS MÁS VENDIDOS", font=('Arial', 14, 'bold'), 
                    bg='#f8f9fa', fg='#2c3e50').pack(pady=(0,20))
            
            # Período
            period_frame = tk.LabelFrame(main_frame, text="Período", font=('Arial', 11), 
                                       bg='#f8f9fa', fg='#2c3e50')
            period_frame.pack(fill='x', pady=(0,15))
            
            period_var = tk.StringVar(value='30dias')
            
            tk.Radiobutton(period_frame, text="Últimos 7 días", variable=period_var, 
                          value='7dias', font=('Arial', 10), bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
            tk.Radiobutton(period_frame, text="Últimos 30 días", variable=period_var, 
                          value='30dias', font=('Arial', 10), bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
            tk.Radiobutton(period_frame, text="Últimos 90 días", variable=period_var, 
                          value='90dias', font=('Arial', 10), bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
            tk.Radiobutton(period_frame, text="Todo el tiempo", variable=period_var, 
                          value='todo', font=('Arial', 10), bg='#f8f9fa').pack(anchor='w', padx=10, pady=2)
            
            def generar_reporte():
                try:
                    # Calcular fechas
                    period = period_var.get()
                    where_clause = ""
                    params = []
                    
                    if period != 'todo':
                        dias = int(period.replace('dias', ''))
                        fecha_inicio = datetime.now() - timedelta(days=dias)
                        where_clause = "WHERE o.fecha_orden >= %s"
                        params.append(fecha_inicio)
                    
                    # Query para productos más vendidos
                    query = f"""
                    SELECT 
                        p.id,
                        p.nombre,
                        c.nombre as categoria,
                        SUM(od.cantidad) as total_vendido,
                        SUM(od.subtotal) as ingresos_totales,
                        AVG(od.precio_unitario) as precio_promedio,
                        COUNT(DISTINCT o.id) as ordenes_count
                    FROM productos p
                    JOIN orden_detalles od ON p.id = od.producto_id
                    JOIN ordenes o ON od.orden_id = o.id
                    JOIN categorias c ON p.categoria_id = c.id
                    {where_clause}
                    GROUP BY p.id, p.nombre, c.nombre
                    ORDER BY total_vendido DESC
                    LIMIT 50
                    """
                    
                    productos = db.execute_query(query, params)
                    
                    if not productos:
                        messagebox.showinfo("Información", "No hay datos de ventas en el período seleccionado")
                        return
                    
                    # Crear Excel
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title = "Productos Más Vendidos"
                    
                    # Título del reporte
                    periodo_text = {
                        '7dias': 'Últimos 7 días',
                        '30dias': 'Últimos 30 días', 
                        '90dias': 'Últimos 90 días',
                        'todo': 'Todo el tiempo'
                    }[period]
                    
                    sheet.merge_cells('A1:G1')
                    title_cell = sheet['A1']
                    title_cell.value = f"PRODUCTOS MÁS VENDIDOS - {periodo_text}"
                    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
                    title_cell.fill = PatternFill('solid', start_color='2980B9')
                    title_cell.alignment = Alignment(horizontal='center')
                    
                    # Headers
                    headers = ['Rank', 'Producto', 'Categoría', 'Cantidad Vendida', 'Ingresos', 'Precio Prom.', 'Órdenes']
                    for col, header in enumerate(headers, 1):
                        cell = sheet.cell(row=3, column=col, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill('solid', start_color='3498DB')
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Datos
                    for rank, producto in enumerate(productos, 1):
                        row = rank + 3
                        sheet.cell(row=row, column=1, value=rank)
                        sheet.cell(row=row, column=2, value=producto['nombre'])
                        sheet.cell(row=row, column=3, value=producto['categoria'])
                        sheet.cell(row=row, column=4, value=producto['total_vendido'])
                        sheet.cell(row=row, column=5, value=float(producto['ingresos_totales']))
                        sheet.cell(row=row, column=6, value=float(producto['precio_promedio']))
                        sheet.cell(row=row, column=7, value=producto['ordenes_count'])
                        
                        # Formatear números
                        sheet.cell(row=row, column=5).number_format = '$#,##0.00'
                        sheet.cell(row=row, column=6).number_format = '$#,##0.00'
                    
                    # Ajustar columnas
                    sheet.column_dimensions['A'].width = 8
                    sheet.column_dimensions['B'].width = 25
                    sheet.column_dimensions['C'].width = 18
                    sheet.column_dimensions['D'].width = 12
                    sheet.column_dimensions['E'].width = 12
                    sheet.column_dimensions['F'].width = 12
                    sheet.column_dimensions['G'].width = 10
                    
                    # Guardar archivo
                    filename = f"productos_vendidos_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    filepath = f"/mnt/user-data/outputs/{filename}"
                    wb.save(filepath)
                    
                    dialog.destroy()
                    messagebox.showinfo("Éxito", f"Reporte generado correctamente\nArchivo: {filename}")
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Error generando reporte: {str(e)}")
            
            # Botones
            button_frame = tk.Frame(main_frame, bg='#f8f9fa')
            button_frame.pack(fill='x', pady=(15,0))
            
            tk.Button(button_frame, text="CANCELAR", font=('Arial', 11, 'bold'),
                     bg='#95a5a6', fg='white', width=12, height=1,
                     command=dialog.destroy).pack(side='left', padx=5)
            
            tk.Button(button_frame, text="GENERAR REPORTE", font=('Arial', 11, 'bold'),
                     bg='#3498db', fg='white', width=15, height=1,
                     command=generar_reporte).pack(side='right', padx=5)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error iniciando reporte: {str(e)}")
    
    def reporte_analisis_precios(self):
        """Análisis de precios por categoría"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener datos
            query = """
            SELECT 
                c.nombre as categoria,
                COUNT(p.id) as total_productos,
                MIN(p.precio) as precio_min,
                MAX(p.precio) as precio_max,
                AVG(p.precio) as precio_promedio,
                SUM(CASE WHEN p.activo = 1 THEN 1 ELSE 0 END) as productos_activos
            FROM categorias c
            LEFT JOIN productos p ON c.id = p.categoria_id
            WHERE c.activo = 1
            GROUP BY c.id, c.nombre
            ORDER BY c.nombre
            """
            
            data = db.execute_query(query)
            
            # Crear Excel
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Análisis de Precios"
            
            # Título
            sheet.merge_cells('A1:F1')
            title_cell = sheet['A1']
            title_cell.value = f"ANÁLISIS DE PRECIOS POR CATEGORÍA - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill('solid', start_color='2980B9')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['Categoría', 'Total Productos', 'Precio Mínimo', 'Precio Máximo', 'Precio Promedio', 'Activos']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, item in enumerate(data, 4):
                sheet.cell(row=row, column=1, value=item['categoria'])
                sheet.cell(row=row, column=2, value=item['total_productos'] or 0)
                sheet.cell(row=row, column=3, value=float(item['precio_min']) if item['precio_min'] else 0)
                sheet.cell(row=row, column=4, value=float(item['precio_max']) if item['precio_max'] else 0)
                sheet.cell(row=row, column=5, value=float(item['precio_promedio']) if item['precio_promedio'] else 0)
                sheet.cell(row=row, column=6, value=item['productos_activos'] or 0)
                
                # Formatear precios
                for col in [3, 4, 5]:
                    sheet.cell(row=row, column=col).number_format = '$#,##0.00'
            
            # Ajustar columnas
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                sheet.column_dimensions[col].width = 18
            
            # Guardar
            filename = f"analisis_precios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Análisis de precios generado\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando análisis: {str(e)}")
    
    def reporte_inventario_completo(self):
        """Reporte completo del inventario"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener todos los productos con información detallada
            query = """
            SELECT 
                p.id,
                p.nombre,
                c.nombre as categoria,
                p.precio,
                CASE WHEN p.activo = 1 THEN 'Activo' ELSE 'Inactivo' END as estado,
                COALESCE(ventas.total_vendido, 0) as total_vendido,
                COALESCE(ventas.ingresos, 0) as ingresos_totales
            FROM productos p
            JOIN categorias c ON p.categoria_id = c.id
            LEFT JOIN (
                SELECT 
                    od.producto_id,
                    SUM(od.cantidad) as total_vendido,
                    SUM(od.subtotal) as ingresos
                FROM orden_detalles od
                GROUP BY od.producto_id
            ) ventas ON p.id = ventas.producto_id
            ORDER BY c.nombre, p.nombre
            """
            
            productos = db.execute_query(query)
            
            # Crear Excel
            wb = Workbook()
            
            # Hoja de productos
            sheet_productos = wb.active
            sheet_productos.title = "Inventario Completo"
            
            # Título
            sheet_productos.merge_cells('A1:G1')
            title_cell = sheet_productos['A1']
            title_cell.value = f"INVENTARIO COMPLETO - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill('solid', start_color='2980B9')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['ID', 'Nombre', 'Categoría', 'Precio', 'Estado', 'Vendido', 'Ingresos']
            for col, header in enumerate(headers, 1):
                cell = sheet_productos.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, producto in enumerate(productos, 4):
                sheet_productos.cell(row=row, column=1, value=producto['id'])
                sheet_productos.cell(row=row, column=2, value=producto['nombre'])
                sheet_productos.cell(row=row, column=3, value=producto['categoria'])
                sheet_productos.cell(row=row, column=4, value=float(producto['precio']))
                sheet_productos.cell(row=row, column=5, value=producto['estado'])
                sheet_productos.cell(row=row, column=6, value=producto['total_vendido'])
                sheet_productos.cell(row=row, column=7, value=float(producto['ingresos_totales']))
                
                # Formatear precios
                sheet_productos.cell(row=row, column=4).number_format = '$#,##0.00'
                sheet_productos.cell(row=row, column=7).number_format = '$#,##0.00'
            
            # Ajustar columnas
            sheet_productos.column_dimensions['A'].width = 8
            sheet_productos.column_dimensions['B'].width = 25
            sheet_productos.column_dimensions['C'].width = 18
            sheet_productos.column_dimensions['D'].width = 12
            sheet_productos.column_dimensions['E'].width = 12
            sheet_productos.column_dimensions['F'].width = 12
            sheet_productos.column_dimensions['G'].width = 12
            
            # Guardar
            filename = f"inventario_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Inventario completo generado\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando inventario: {str(e)}")
    
    def reporte_categorias_resumen(self):
        """Resumen de categorías con estadísticas"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Obtener datos de categorías con estadísticas
            query = """
            SELECT 
                c.id,
                c.nombre,
                CASE WHEN c.activo = 1 THEN 'Activa' ELSE 'Inactiva' END as estado,
                COUNT(p.id) as total_productos,
                SUM(CASE WHEN p.activo = 1 THEN 1 ELSE 0 END) as productos_activos,
                MIN(p.precio) as precio_min,
                MAX(p.precio) as precio_max,
                AVG(p.precio) as precio_promedio
            FROM categorias c
            LEFT JOIN productos p ON c.id = p.categoria_id
            GROUP BY c.id, c.nombre, c.activo
            ORDER BY c.nombre
            """
            
            categorias = db.execute_query(query)
            
            # Crear Excel
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Resumen Categorías"
            
            # Título
            sheet.merge_cells('A1:H1')
            title_cell = sheet['A1']
            title_cell.value = f"RESUMEN DE CATEGORÍAS - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill('solid', start_color='2980B9')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['ID', 'Categoría', 'Estado', 'Total Prod.', 'Activos', 'P. Mínimo', 'P. Máximo', 'P. Promedio']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='3498DB')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row, categoria in enumerate(categorias, 4):
                sheet.cell(row=row, column=1, value=categoria['id'])
                sheet.cell(row=row, column=2, value=categoria['nombre'])
                sheet.cell(row=row, column=3, value=categoria['estado'])
                sheet.cell(row=row, column=4, value=categoria['total_productos'] or 0)
                sheet.cell(row=row, column=5, value=categoria['productos_activos'] or 0)
                sheet.cell(row=row, column=6, value=float(categoria['precio_min']) if categoria['precio_min'] else 0)
                sheet.cell(row=row, column=7, value=float(categoria['precio_max']) if categoria['precio_max'] else 0)
                sheet.cell(row=row, column=8, value=float(categoria['precio_promedio']) if categoria['precio_promedio'] else 0)
                
                # Formatear precios
                for col in [6, 7, 8]:
                    sheet.cell(row=row, column=col).number_format = '$#,##0.00'
            
            # Ajustar columnas
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 10
            sheet.column_dimensions['F'].width = 12
            sheet.column_dimensions['G'].width = 12
            sheet.column_dimensions['H'].width = 12
            
            # Guardar
            filename = f"categorias_resumen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/mnt/user-data/outputs/{filename}"
            wb.save(filepath)
            
            messagebox.showinfo("Éxito", f"Resumen de categorías generado\nArchivo: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando resumen: {str(e)}")

    def logout(self):
        if auth.current_user and auth.current_user['tipo'] == 'cajero' and auth.current_turno:
            response = messagebox.askyesno(
                "Cerrar Sesión", 
                "Tiene un turno abierto. ¿Desea cerrar sesión sin cerrar el turno?\n\n"
                "Recuerde que deberá volver para cerrar su turno al final del día."
            )
            if not response:
                return
        
        auth.logout()
        self.show_login()

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = RestaurantPOS()
    app.run()
